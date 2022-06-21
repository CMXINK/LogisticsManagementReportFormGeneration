import re
import sys
import datetime

import openpyxl
from PyQt5.QtGui import QWheelEvent, QIcon
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from resource.ui.usr_jizhuang import Ui_Form
from mainauto_read_wite_file import form_operation
import os
from mysql import Mysql
import myopne_icon_rc


class User_jizhuang(Ui_Form, QWidget):
    textedit_finished = pyqtSignal()
    remember_history = pyqtSignal(str, int, str, str, str, list, dict)

    def __init__(self, g_User, g_Pw, g_Port, g_Ip, username, password, parent=None, *args, **kwargs):
        super(User_jizhuang, self).__init__(parent)
        self.setWindowIcon(QIcon(":/open.ico"))
        self.history_data = {}
        self.g_User = g_User
        self.g_Pw = g_Pw
        self.g_Port = g_Port
        self.g_Ip = g_Ip
        self.user = username
        self.password = password
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setupUi(self)
        self.object_name_list = []
        self.listwidget.hide()
        self.tidanhao_lineedit.setFocus()
        self.operation_help_btn.clicked.connect(self.help_lk)
        self.setWindowTitle("物流管理系统")
        t = 1
        self.combobox_data = []
        try:
            while True:
                self.combobox_data.append(self.operation_from_combobox.itemText(t).strip())
                if not self.operation_from_combobox.itemText(t):
                    break
                t += 1
        except:
            pass

    def fom_lk(self, index):
        global save_name, okPressed, filename
        print(index)
        if index != 0:
            filename = self.operation_from_combobox.itemText(index).strip()
            save_name, okPressed = QInputDialog.getText(self, filename, "请输入保存后的文件名称:", QLineEdit.Normal, "")
        if save_name or okPressed:
            data = []
            self.history_data[filename] = save_name
            if index == 1:
                data.append(self.tidanhao_lineedit.text())
                # 预计开船日存在格式问题
                data.append(self.yujikaichuanri_lineedit.text())
                data.append(self.chuanming_lineedit.text() + '\n' + self.hangci_lineedit.text())
                data.append(self.zhuanghuogang_lineedit.text())
                data.append(self.mudidi_lineedit.text())
                data.append(self.xiangxing_lineedit.text() + "*" + self.xiangliang_lineedit.text())
                data.append(datetime.date.today().strftime("%Y/%m/%d"))
            if index == 2:
                data.append(self.fapiaotaitou_lineedit.text())
                data.append(self.chuanming_lineedit.text() + "  " + self.hangci_lineedit.text())
                data.append(self.tidanhao_lineedit.text())
                data.append(self.zhuanghuogang_lineedit.text())
                data.append(self.mudidi_lineedit.text())
                data.append(self.xiangxing_lineedit.text() + "*" + self.xiangliang_lineedit.text())
                data.append(datetime.date.today().strftime("%Y/%m/%d"))
            if index == 3:
                data.append(self.tidanhao_lineedit.text())
                # 预计开船日存在格式问题
                data.append(self.yujikaichuanri_lineedit.text())
                data.append(self.chuanming_lineedit.text() + '  ' + self.hangci_lineedit.text())
                data.append(self.zhuanghuogang_lineedit.text())
                data.append(self.mudidi_lineedit.text())
                data.append(self.xiangxing_lineedit.text() + "*" + self.xiangliang_lineedit.text())
                data.append(self.xiangxing_lineedit.text())
                data.append(self.jianshu_lineedit.text())
                data.append(datetime.date.today().strftime("%Y/%m/%d"))
            if index == 4:
                data = [self.chuanming_lineedit.text(), self.tidanhao_lineedit.text(), self.mudidi_lineedit.text(),
                        self.pinming_lineedit.text(),
                        self.xiangxing_lineedit.text() + "*" + self.xiangliang_lineedit.text(),
                        datetime.date.today().strftime("%Y-%m-%d")]
                form_operation().jizhuangxiang_docx_file(filename, save_name, data)
                self.operation_from_combobox.setCurrentIndex(0)
            if index == 5:
                data.append(self.fahuoren_textedit.toPlainText())
                data.append(self.tidanhao_lineedit.text())
                group = [self.operation_genggai_radio, self.operation_fenpaio_radio, self.operation_dianfang_radio,
                         self.operation_zhengwen_radio, self.operation_hepiao_radio]
                btn_data = ""
                for radio in group:
                    if radio.isChecked():
                        btn_data = radio.text()
                data.append(btn_data)

                data.append(self.shouhuoren_textedit.toPlainText())
                data.append(self.tongzhiren_textedit.toPlainText())

                data.append(self.chuanming_lineedit.text() + "  " + self.hangci_lineedit.text())
                data.append(self.zhuanghuogang_lineedit.text())
                data.append(self.xiehuogang_lineedit.text())
                data.append(self.mudidi_lineedit.text())
                data.append(self.maitou_textedit.toPlainText())
                data.append(self.jianshu_lineedit.text())
                data.append(self.huomiao_textedit.toPlainText())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.tiji_lineedit_2.text())
            if index == 6:
                data.append(self.yujikaichuanri_lineedit.text())
                data.append(self.weituodanwei_lineedit.text())
                data.append(self.tidanhao_lineedit.text())
                data.append(self.xianghao_lineedit.text())
                data.append(self.qianfenghao_lineedit.text())
                data.append(self.pinming_lineedit.text())
                data.append(self.jianshu_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.tiji_lineedit_2.text())
                data.append(self.mudidi_lineedit.text())
                data.append(self.yewuyuan_lineedit.text())
                data.append(self.dailidianhua_lineedit.text())
            if index != 4 and index > 0:
                form_operation().copy_jizhuangxiang_file(filename, save_name, data)
                self.operation_from_combobox.setCurrentIndex(0)
        else:
            self.operation_from_combobox.setCurrentIndex(0)

    def fresh_lk(self):
        data = Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).check_oper(grade=3, text=self.text)
        data = data[0]
        self.creat_new(self.text, data)

    def creat_new(self, text, data):
        self.text = text
        self.ship = data[0]
        data0 = data[:4]
        self.qiyungang = data[4]
        data1 = data[5:18]
        data2 = data[20:-1]
        self.user = data[-1]
        data = data0 + data1 + data2
        list1 = [self.chuanming_lineedit, self.hangci_lineedit, self.tidanhao_lineedit,
                 self.yujikaichuanri_lineedit, self.zhuanghuogang_lineedit, self.xiehuogang_lineedit,
                 self.mudidi_lineedit, self.xiangxing_lineedit, self.xiangliang_lineedit,
                 self.xianghao_lineedit, self.pinming_lineedit, self.fapiaotaitou_lineedit,
                 self.zhibiaoriqi_lineedit, self.qianfenghao_lineedit, self.weituodanwei_lineedit,
                 self.yewuyuan_lineedit, self.dailidianhua_lineedit,
                 self.huoming_lineedit, self.jianshu_lineedit, self.zhongliang_lineedit,
                 self.tiji_lineedit_2, self.jifeidun_lineedit, self.fahuoren_textedit,
                 self.shouhuoren_textedit, self.tongzhiren_textedit, self.maitou_textedit,
                 self.huomiao_textedit]
        for i in range(0, len(list1)):

            if i in self.object_name_list:
                list1[i].disconnect()
                list1[i].setText(data[i])
                list1[i].installEventFilter(self)
                list1[i].textChanged.connect(lambda: self.change_listwidget(list1[i].objectName()))
                if isinstance(i, QLineEdit):
                    list1[i].editingFinished.connect(lambda: self.write_file(list1[i].objectName()))
            else:
                list1[i].setText(data[i])

        self.conn(self.jianshu_lineedit)
        self.conn(self.chuanming_lineedit)
        self.conn(self.hangci_lineedit)
        self.conn(self.zhuanghuogang_lineedit)
        self.conn(self.xiehuogang_lineedit)
        self.conn(self.xiangxing_lineedit)
        self.conn(self.pinming_lineedit)
        self.conn(self.fapiaotaitou_lineedit)
        self.conn(self.weituodanwei_lineedit)
        self.conn(self.yewuyuan_lineedit)
        self.conn(self.dailidianhua_lineedit)
        self.conn(self.zhongliang_lineedit)
        self.conn(self.tidanhao_lineedit)
        self.conn(self.jifeidun_lineedit)
        self.conn(self.fahuoren_textedit)
        self.conn(self.shouhuoren_textedit)
        self.conn(self.huoming_lineedit)
        self.conn(self.maitou_textedit)
        self.conn(self.tongzhiren_textedit)

    def reset_lk(self):
        list1 = [self.tidanhao_lineedit, self.yujikaichuanri_lineedit, self.chuanming_lineedit,
                 self.hangci_lineedit, self.zhuanghuogang_lineedit, self.xiehuogang_lineedit,
                 self.mudidi_lineedit, self.xiangxing_lineedit, self.xiangliang_lineedit,
                 self.xianghao_lineedit, self.pinming_lineedit, self.fapiaotaitou_lineedit,
                 self.zhibiaoriqi_lineedit, self.qianfenghao_lineedit, self.weituodanwei_lineedit,
                 self.yewuyuan_lineedit, self.dailidianhua_lineedit, self.beizhu_lineedit,
                 self.huoming_lineedit, self.jianshu_lineedit, self.zhongliang_lineedit,
                 self.tiji_lineedit_2, self.jifeidun_lineedit, self.fahuoren_textedit,
                 self.shouhuoren_textedit, self.tongzhiren_textedit, self.maitou_textedit,
                 self.huomiao_textedit]
        for i in list1:

            if i in self.object_name_list:
                i.disconnect()
                i.setText("")
                i.installEventFilter(self)
                i.textChanged.connect(lambda: self.change_listwidget(i.objectName()))
                if isinstance(i, QLineEdit):
                    i.editingFinished.connect(lambda: self.write_file(i.objectName()))
            else:
                i.setText("")
        for i in [self.operation_fahuo_checkbox, self.operation_daohuo_checkbox,
                  self.operation_dizai_checkbox, self.operation_jiancha_checkbox,
                  self.operation_fangxing_checkbox, self.operation_zhuangzai_chexkbox]:
            i.setChecked(False)
        for i in [self.operation_zhengwen_radio, self.operation_dianfang_radio, self.operation_genggai_radio,
                  self.operation_hepiao_radio, self.operation_fenpaio_radio]:
            i.setCheckable(False)
            i.setCheckable(True)
            i.setChecked(False)

    def SQL_lk(self):
        if not self.tidanhao_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>提单号为空</strong>")
            return None
        if not self.chuanming_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>船名为空</strong>")
            return None
        if self.chuanming_lineedit.text() == self.tidanhao_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>船名不应该和提单号相同</strong>")
            return None
        list1 = [self.chuanming_lineedit, self.hangci_lineedit, self.tidanhao_lineedit,
                 self.yujikaichuanri_lineedit, self.zhuanghuogang_lineedit, self.xiehuogang_lineedit,
                 self.mudidi_lineedit, self.xiangxing_lineedit, self.xiangliang_lineedit,
                 self.xianghao_lineedit, self.pinming_lineedit, self.fapiaotaitou_lineedit,
                 self.zhibiaoriqi_lineedit, self.qianfenghao_lineedit, self.weituodanwei_lineedit,
                 self.yewuyuan_lineedit, self.dailidianhua_lineedit,
                 self.huoming_lineedit, self.jianshu_lineedit, self.zhongliang_lineedit,
                 self.tiji_lineedit_2, self.jifeidun_lineedit, self.fahuoren_textedit,
                 self.shouhuoren_textedit, self.tongzhiren_textedit, self.maitou_textedit,
                 self.huomiao_textedit]
        data = []
        for i in list1:
            try:
                if isinstance(i, QLineEdit):
                    data.append(i.text())
                elif isinstance(i, QTextEdit):
                    data.append(i.toPlainText())
            except:
                data.append("")
        data1 = data[:17]
        data2 = data[17:]
        temp = ""
        try:
            for i in [self.operation_fahuo_checkbox, self.operation_daohuo_checkbox,
                      self.operation_dizai_checkbox, self.operation_jiancha_checkbox,
                      self.operation_fangxing_checkbox, self.operation_zhuangzai_chexkbox]:
                if i.isChecked():
                    temp = i.text()
        except:
            temp = ""
        data1.append(temp)
        temp1 = ""
        for i in [self.operation_zhengwen_radio, self.operation_dianfang_radio, self.operation_genggai_radio,
                  self.operation_hepiao_radio, self.operation_fenpaio_radio]:
            if i.isChecked():
                temp1 = i.text()
        data1.append(temp1)
        data = data1 + data2
        data.append(self.user)
        # temp_data_no_qiyungang = data[:4]
        # temp_data_no_qiyungang.append(self.qiyungang)
        # data = temp_data_no_qiyungang + data[4:]
        print("data====data=====data=", data)
        Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).save_jizhuangxiang_oper(self.text, data)

    def help_lk(self):
        self.help_content = '<p style="line-height: 25px; font-size: 16px" ><b>下货纸 ' \
                            ':</b>下货纸存在位置是当前软件同级目录下的Excel_files,如果需要新增模板，请将模板添加到template''目录下,且模板需与之前模板所需内容相似<br>''<p ' \
                            'style="line-height: 25px; font-size: 16px" ><b>刷新: ' \
                            '</b>点击刷新会重新从数据库提取数据中所用拥有的数据到当前界面<br>''<p style="line-height: 25px; font-size: 16px" ' \
                            '><b>SQL: </b>会将当前数据保存到数据库,既保存<br>''<p style="line-height: 25px; font-size: 16px" ><b>重置: ' \
                            '</b>点击重置会清空当前界面输入框所有内容<br>''<p style="line-height: 25px; font-size: 16px" ><b>其他: ' \
                            '</b>向上滑动滚轮界面将上移,向下滑动滚轮界面将下移<br>' \
                            '<p style="line-height: 25px; font-size: 16px"; font-color:red ><b>注意 ' \
                            ':</b>由于软件需要,请操作完后点击退出保证软件正常运行'
        QMessageBox.information(self, '帮助', self.help_content)

    def conn(self, object_name):
        """
        :param object_name: lineedit对象
        :param listwidget: listwidget对象
        :return: none
        """
        self.object_name_list.append(object_name)
        object_name.installEventFilter(self)
        self.listwidget.hide()
        self.listwidget.clicked.connect(self.lineedit_set_text)

        object_name.textChanged.connect(lambda: self.change_listwidget(object_name.objectName()))
        if isinstance(object_name, QLineEdit):
            object_name.editingFinished.connect(lambda: self.write_file(object_name.objectName()))
        # 创建数据存储的文件
        self.listwidget_data = {}
        self.remember_data = {}
        row_data = []

        self.file = os.path.abspath(".") + "\\resource\\" + 'listwidget_data.xlsx'
        if os.path.exists(self.file):
            wb = openpyxl.load_workbook(filename=self.file)
            sheets = wb.sheetnames
            wb.close()
            if object_name.objectName() not in sheets:
                wb = openpyxl.load_workbook(filename=self.file)
                wb.create_sheet(object_name.objectName())
                wb.save(self.file)
                wb.close()
        else:
            wb = openpyxl.Workbook()
            wb.create_sheet(object_name.objectName())
            wb.save(self.file)
            wb.close()
        wb = openpyxl.load_workbook(filename=self.file, read_only=True)
        sheets = wb.sheetnames
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        row_data.append(cell.value)
            self.listwidget_data[sheet_name] = row_data
            self.remember_data[sheet_name] = []
            row_data = []
        self.current_listdata = []
        wb.close()

    def my_re(self, str):
        if isinstance(self.focuse_widget, QTextEdit):
            target = self.focuse_widget.toPlainText()
        else:
            target = self.focuse_widget.text()
        """
        :param traget: 目标字符串
        :param str: 查找的字符串
        :return: 前几位匹配的字符串
        """
        if re.match(target, str, re.I):
            return str

    #
    # 当lineedit控件输入框内容变化时触发，查找所有历史输入若存在首位相同显示listwidget控件
    # 并向中插入匹配的结果作为item.每次显示时清空上次items项
    def change_listwidget(self, objectname):
        objectname = self.focuse_widget.objectName()
        if self.listwidget_data.get(objectname):
            self.current_listdata = list(
                filter(lambda t: t is not None, list(map(self.my_re, self.listwidget_data.get(objectname)))))
            print('self.current_listwidget=', self.current_listdata)
            if self.current_listdata:
                self.listwidget.disconnect()
                self.listwidget.clear()
                self.listwidget.clicked.connect(self.lineedit_set_text)
                self.listwidget.addItems(sorted(self.current_listdata))
                self.listwidget.show()
        if isinstance(self.focuse_widget, QTextEdit):
            if self.focuse_widget.toPlainText() and not self.current_listdata:
                print(self.current_listdata)
                self.listwidget.hide()
        elif self.focuse_widget.text() and not self.current_listdata:
            print(self.current_listdata)
            self.listwidget.hide()

        if not self.listwidget.count():
            self.listwidget.hide()

    # 当点击listwidget的item时触发,将点击的item内容设置为当前lineedit文本
    def lineedit_set_text(self, index):
        row = index.row()
        self.focuse_widget.setText(self.current_listdata[row])

    # 编辑完成时触发，将lineedit文本内容与所有历史输入内容对比，不同则记住作为本次新增的输入内容
    def write_file(self, object):
        object = self.focuse_widget
        objectname = self.focuse_widget.objectName()
        if isinstance(object, QLineEdit):
            text = self.focuse_widget.text()
        else:
            text = object.toPlainText()
        try:
            if text in self.listwidget_data.get(objectname):
                pass
            else:
                if text:
                    self.remember_data[objectname].append(text)
                    self.listwidget_data[objectname].append(text)
        except:
            if text:
                self.remember_data[objectname].append(text)
                self.listwidget_data[objectname].append(text)

    # 当关闭程序时触发，将本次记住的新增的输入内容写入文件
    def closeEvent(self, event):
        showdata = ""
        count = 1  # 用来计数
        if "" in self.combobox_data:
            self.combobox_data = self.combobox_data[:-1]
        for i in self.combobox_data:
            try:
                if self.history_data.get(i):
                    count -= 1
                    pass
                else:
                    self.history_data[i] = ""
                    showdata += str(count) + ": " + i + "<br>"
            except:
                self.history_data[i] = ""
                showdata += str(count) + ": " + i + "<br>"
            count += 1
        # if showdata:
        #     QMessageBox.about(self, '未完成操作', showdata)
        if showdata:
            selectdata = [i[3:] for i in showdata.split("<br>")[:-1]]
            selectdata.insert(0, {"tidanhao_num": self.tidanhao_lineedit.text()})
            res_data = Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).select_if_complate_data(selectdata, grade='3')
            print("===============>",res_data,"=============>")
            if res_data != "all_exists":
                QMessageBox.about(self, '未完成操作', res_data if res_data else showdata)
        reply = QMessageBox.question(self, "提问", "确定退出吗?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.history_data["time"] = datetime.datetime.now().strftime("%Y-%m-%d-%H:%M.%S")
            self.remember_history.emit(self.user, 3, self.ship, self.text, self.tidanhao_lineedit.text(),
                                       self.combobox_data,
                                       self.history_data)  # 船名， 提单号，combobox数据， 记录数据
            # 把时间存入数据表中
            Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).save_jizhuangxiang_time(self.text,
                                                                                          self.tidanhao_lineedit.text(),
                                                                                          self.combobox_data,
                                                                                          self.history_data)
            self.history_data = {}

            wb = openpyxl.load_workbook(self.file)
            sheets = wb.sheetnames
            for sheet_name in sheets:
                sheet = wb[sheet_name]
                if self.remember_data.get(sheet_name):
                    sheet.append(self.remember_data.get(sheet_name))
            wb.save(self.file)
            wb.close()

            event.accept()
        if reply == QMessageBox.No:
            event.ignore()

    #
    def eventFilter(self, a0: 'QObject', a1: 'QEvent'):
        if a1.type() == QEvent.FocusIn:
            print(a0.objectName())
            self.focuse_widget = a0
            if isinstance(a0, QLineEdit):
                a0.setStyleSheet("border-bottom:none;")
            t = a0
            x = 0
            y = t.height()

            while True:
                x += t.x()
                y += t.y()
                t = t.parent()
                try:
                    check = t.parent().pos()
                except:
                    break
            self.listwidget.move(x, y)
            self.listwidget.disconnect()
            self.listwidget.clear()
            if self.listwidget_data.get(a0.objectName()):
                self.current_listdata = sorted(self.listwidget_data.get(a0.objectName()))
            else:
                self.current_listdata = []
                self.listwidget.hide()
            if self.current_listdata:
                self.listwidget.addItems(self.current_listdata)
            self.listwidget.pressed.connect(self.lineedit_set_text)
            if self.current_listdata:
                self.listwidget.show()
                self.listwidget.raise_()
            self.listwidget.setStyleSheet("QListWidget{"
                                          "border:1px solid rgba(30,30,30,0.3);"
                                          "background-color:rgb(230,230,230);"
                                          "}"

                                          "QListWidget::item"
                                          "{"
                                          "border:1px dashed gray;"
                                          "  }"

                                          "QScrollBar"
                                          "{"
                                          "width:px;"
                                          "background:rgb(0,0,0,0%);"
                                          "margin:0px,0px,0px,0px;"
                                          "padding-top:0px;"
                                          "padding-bottom:0px;"
                                          "}")
            self.listwidget.setFixedWidth(a0.width())
            self.listwidget.setFixedHeight(100)
        elif a1.type() == QEvent.FocusOut:
            #
            if isinstance(a0, QLineEdit):
                a0.setStyleSheet("border-bottom:1px solid rgb(0,0,0);")
            self.listwidget.move(2000, 2000)
            if isinstance(a0, QTextEdit):
                self.write_file(a0)
        return QWidget.eventFilter(self, a0, a1)

    def wheelEvent(self, event):
        if self.mapFromGlobal(self.cursor().pos()).x() in range(self.listwidget.pos().x(),
                                                                self.listwidget.pos().x() + self.listwidget.width()) and self.mapFromGlobal(
            self.cursor().pos()).y() in range(self.listwidget.pos().y(),
                                              self.listwidget.pos().y() + self.listwidget.height()) and self.listwidget.isVisible():
            pass
        elif QWheelEvent.angleDelta(event).y() > 0:
            self.move(self.pos().x(), self.pos().y() + 15)
        elif QWheelEvent.angleDelta(event).y() < 0:
            self.move(self.pos().x(), self.pos().y() - 15)

    def keyPressEvent(self, event):
        if event.key() == 16777220:
            try:
                if isinstance(self.focuse_widget, QLineEdit):
                    self.listwidget.hide()
            except:
                pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main = User_jizhuang()
    main.show()
    app.exit(app.exec_())
