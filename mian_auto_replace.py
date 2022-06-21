import sys
from functools import wraps
from PyQt5.Qt import *
from PyQt5.uic.properties import QtCore
import re
import openpyxl
from mysql import Mysql
from resource.ui.main_auto import Ui_Form
from constant import Con_main_auto as constant
from mainauto_read_wite_file import Read_file as Read, Write_file as Write
import math
import myopne_icon_rc
from threading import Thread


class Main_auto(QWidget, Ui_Form):
    main_change_account_number_signal = pyqtSignal()
    oper_signal = pyqtSignal(str, str, str)
    tips_signal = pyqtSignal(int)  # grade, user, data

    # tips_real_signal = pyqtSignal()  # 真正的tips
    # g_User, g_Pw, g_Port, g_Ip, sanhuo_tips, jizhuangxiang_tips,user_sanhuo, user_jizhuangxiang,check_db_no_conn, check_user_no_conn
    def __init__(self, g_User, g_Pw, g_Port, g_Ip, sanhuo_tips, jizhuangxiang_tips, user_sanhuo, user_jizhuangxiang,
                 check_db_no_conn=True,
                 check_user_no_conn=True, parent=None, *args,
                 **kwargs):
        super(Main_auto, self).__init__(parent)
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setupUi(self)
        self.setWindowIcon(QIcon(":/open.ico"))
        self.mainauto_init_style()
        self.setWindowTitle("物流管理系统")
        self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
        self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)
        self.treewidget.itemClicked.connect(self.treewidget_temp)
        self.listWidget.itemClicked.connect(self.listwidget_temp)
        # self.treewidget.itemDoubleClicked(self.mousePressEvent)
        # 初始化或初始定义数据
        self.undo_data = []
        self.match_result = 0
        self.undo_redo_count = 0
        self.old_all_undo_stept = 0
        # 用户数据初始化
        self.oper_user = ""
        self.oper_password = ""
        # 数据库的数据初始化
        self.user = g_User
        self.password = g_Pw
        self.port = int(g_Port)
        self.ip = g_Ip
        self.sanhuo_tips = sanhuo_tips
        self.jizhuangxiang_tips = jizhuangxiang_tips
        self.user_sanhuo = user_sanhuo
        self.user_jizhuangxiang = user_jizhuangxiang
        self.listWidget.hide()
        #  拿到现在用户操作的状态(是否连接数据库,是否登录)
        """
        测试数据
        """

        self.check_db_no_conn = check_db_no_conn
        self.check_user_no_conn = check_user_no_conn
        self.if_sql_init(self.check_db_no_conn, self.check_user_no_conn, self.oper_user, self.oper_password)

    def treewidget_temp(self, object):
        # x = QCursor.pos().x() - self.widget.pos().x()
        # y = QCursor.pos().y() - self.widget.pos().y()
        # self.listWidget.setParent(self.treewidget.parent())
        # self.listWidget.hide()
        ##f2f2f2
        self.listWidget.hide()
        if object.parent() == self.root1 or object.parent() == self.root2 or object.parent() == self.root3:
            if self.grade != 1:
                self.listWidget.resize(60, 80)
                self.listWidget.setStyleSheet("QListWidget{"
                                              " background-color:white;"
                                              "height:60px;"
                                              "width:60px;"
                                              "border:2px solid #f2f2f2;"

                                              "font-size:15px;"
                                              "text-align:center"
                                              "}"

                                              "QListWidget::item"
                                              "{"
                                              "height:15px;"
                                              "border:none;"
                                              "padding:5px;"
                                              "}"
                                              "QListWidget::item::hover"
                                              "{"
                                              "background-color:#91c9f7;"
                                              "border:2px solid #f2f2f2;"
                                              "}"
                                              "QListWidget::item:"
                                              "{"
                                              "background-color:#91c9f7;"
                                              "border:2px solid #f2f2f2;}"
                                              )
                self.listWidget.move(QCursor.pos().x() - self.window().geometry().x() + 7,
                                     QCursor.pos().y() - self.window().geometry().y())
                self.listWidget.setFocusPolicy(Qt.NoFocus)
                self.listWidget.show()
                self.re_object_name = object.text(0)
        self.oper_signal.emit(object.text(0), str(self.grade), self.oper_user)

    def listwidget_temp(self, object):
        if object.text() == "添加":
            self.listWidget.hide()
            ship = self.re_object_name.split(" &")[0]
            ship_num = self.re_object_name.split(" &")[1]
            tidanhao = self.get_check_tidanhao(ship, ship_num)
            print("==============>tiandanhao", tidanhao)
            if tidanhao:
                #     调用增一行,达到新建行的目的
                self.add_one_row(ship, ship_num, tidanhao)
        elif object.text() == "删除":
            self.listWidget.hide()
        elif object.text() == "取消":
            self.listWidget.hide()

    # 用于主页左边新建提对提单号进行检测
    def get_check_tidanhao(self, ship, ship_num):
        tidanhao, okPressed = QInputDialog.getText(self, "提单号",
                                                   "请输入{}".format(ship) + " ==> " + "{}的新提单号:".format(ship_num),
                                                   QLineEdit.Normal, "")
        if tidanhao or okPressed:
            if tidanhao:
                list_tidanhao = []
                for i in range(0, self.all_rows):
                    if self.tableWidget.item(i, 4).text():
                        list_tidanhao.append(self.tableWidget.item(i, 4).text())
                if tidanhao in list_tidanhao:
                    reply = QMessageBox.warning(self, '警告',
                                                "<strong>提单号已存在</strong>,请核对后再重新填写",
                                                QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        return self.get_check_tidanhao(ship, ship_num)
                    else:
                        return False
                else:
                    return tidanhao
            else:
                reply = QMessageBox.warning(self, '警告',
                                            "<strong>提单号不能为空</strong>,请重新填写",
                                            QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    return self.get_check_tidanhao(ship, ship_num)
                else:
                    return False

    def if_sql_init(self, check_db_no_conn, check_user_no_conn, oper_user, oper_password, user='root', password='root',
                    port=3306, ip=''):
        global type_user, root, data_staff
        self.oper_user = oper_user
        self.oper_password = oper_password
        self.user = user
        self.password = password
        self.port = int(port)
        self.ip = ip
        self.undo_data = []
        self.match_result = 0
        self.undo_redo_count = 0
        self.old_all_undo_stept = 0
        self.check_db_no_conn = check_db_no_conn
        self.check_user_no_conn = check_user_no_conn
        # 当连接数据库时并登陆成功时建立数据表
        self.create_tree()

    def mousePressEvent(self, event):
        if event.buttons() & Qt.LeftButton:
            x = event.x()
            y = event.y()
            if x not in range(self.listWidget.pos().x(), self.listWidget.pos().x() + 60) or y in range(
                    self.listWidget.pos().y(), self.listWidget.pos().y() + 60):
                self.listWidget.hide()

    def create_tree(self):
        user_icon = QIcon("resource\\image\\main_auto_user_icon.png")
        sheep_icon = QIcon("resource\\image\\main_auto_ship_icon.png")
        number_icon = QIcon("resource\\image\\main_auto_number_icon.png")
        staff_icon = QIcon("resource\\image\\main_auto_staff_icon.png")
        if not self.check_db_no_conn and not self.check_user_no_conn and self.oper_user and self.oper_password:
            self.treewidget.clear()
            self.grade = Mysql(self.user, self.password, self.port, self.ip).user_password_table(self.oper_user,
                                                                                                 self.oper_password)
            if self.grade == 1:
                self.mainauto_fucntion_combobox.setItemText(2, "添加")
                self.mainauto_fucntion_combobox.setItemText(5, "更新")
            self.root1 = QTreeWidgetItem(self.treewidget)
            self.root1.setText(0, '散货')
            self.root2 = QTreeWidgetItem(self.treewidget)
            self.root2.setText(0, '集装箱')
            self.root3 = QTreeWidgetItem(self.treewidget)
            self.root3.setText(0, '财务')
            self.root1.setIcon(0, user_icon)
            self.root2.setIcon(0, user_icon)
            self.root3.setIcon(0, user_icon)
            if int(self.grade) == 1:

                self.mainauto_fucntion_combobox.setItemText(2, "添加")
                for i in range(2, 4):
                    if i == 2:
                        root = self.root1
                        type_user = 'user1'
                    elif i == 3:
                        root = self.root2
                        type_user = 'user2'
                    elif i == 4:
                        root = self.root3
                        type_user = 'user3'
                    data_staff = Mysql(self.user, self.password, self.port, self.ip).get_staff_user(i)
                    try:
                        for j in range(0, len(data_staff)):
                            d = QTreeWidgetItem(root)
                            d.setText(0, data_staff[j])
                            d.setIcon(0, staff_icon)
                            data_sheep_num = Mysql(self.user, self.password, self.port, self.ip).get_sheeps_data(
                                type_user,
                                data_staff[
                                    j],
                                self.grade)
                            sheep_num_text = [i[0] + " &" + i[1] for i in data_sheep_num if i[0]]
                            for count_sheeps in range(0, len(sheep_num_text)):
                                t = QTreeWidgetItem(d)
                                t.setText(0, sheep_num_text[count_sheeps])
                                t.setIcon(0, sheep_icon)
                                data_tidanhao = Mysql(self.user, self.password, self.port, self.ip).get_tidanhao_data(
                                    sheep_num_text[count_sheeps].split(" &")[0],
                                    sheep_num_text[count_sheeps].split(" &")[1], type_user, data_staff[j]
                                    , self.grade)
                                for tidanhao in range(0, len(data_tidanhao)):
                                    k = QTreeWidgetItem(t)
                                    k.setText(0, data_tidanhao[tidanhao])
                                    k.setIcon(0, number_icon)
                    except:
                        pass


            else:
                if int(self.grade) == 2:
                    sheep_list = []
                    type_user = "user1"
                    data_sheep_num = Mysql(self.user, self.password, self.port, self.ip).get_sheeps_data(type_user,
                                                                                                         self.oper_user,
                                                                                                         self.grade)
                    sheep_num_text = [i[0] + " &" + i[1] for i in data_sheep_num if i[0]]
                    for i in range(0, len(sheep_num_text)):
                        if sheep_num_text[i] in sheep_list:
                            continue
                        else:
                            sheep_list.append(sheep_num_text[i])
                        t = QTreeWidgetItem(self.root1)
                        t.setIcon(0, sheep_icon)
                        t.setText(0, sheep_num_text[i])

                        data_tidanaho = Mysql(self.user, self.password, self.port, self.ip).get_tidanhao_data(
                            sheep_num_text[i].split(" &")[0], sheep_num_text[i].split(" &")[1],
                            type_user,
                            self.oper_user, self.grade)
                        for index, j in enumerate(data_tidanaho):
                            k = QTreeWidgetItem(t)
                            k.setText(0, j)
                            k.setIcon(0, number_icon)
                elif int(self.grade) == 3:
                    sheep_list = []
                    type_user = "user2"
                    data_sheep_num = Mysql(self.user, self.password, self.port, self.ip).get_sheeps_data(type_user,
                                                                                                         self.oper_user,
                                                                                                         self.grade)
                    sheep_num_text = [i[0] + " &" + i[1] for i in data_sheep_num if i[0]]
                    for i in range(0, len(sheep_num_text)):
                        if sheep_num_text[i] in sheep_list:
                            continue
                        else:
                            sheep_list.append(sheep_num_text[i])
                        t = QTreeWidgetItem(self.root2)
                        t.setIcon(0, sheep_icon)
                        t.setText(0, sheep_num_text[i])

                        data_tidanaho = Mysql(self.user, self.password, self.port, self.ip).get_tidanhao_data(
                            sheep_num_text[i].split(" &")[0], sheep_num_text[i].split(" &")[1],
                            type_user,
                            self.oper_user, self.grade)
                        for j in range(0, len(data_tidanaho)):
                            k = QTreeWidgetItem(t)
                            k.setIcon(0, number_icon)
                            k.setText(0, data_tidanaho[j])
                elif int(self.grade) == 4:
                    sheep_list = []
                    type_user = "user3"
                    data_sheeps = Mysql(self.user, self.password, self.port, self.ip).get_sheeps_data(type_user,
                                                                                                      self.oper_user,
                                                                                                      self.grade)
                    for i in range(0, len(data_sheeps)):
                        if data_sheeps[i] in sheep_list:
                            continue
                        else:
                            sheep_list.append(data_sheeps[i])
                        t = QTreeWidgetItem(self.root3)
                        t.setIcon(0, sheep_icon)
                        t.setText(0, data_sheeps[i])

                        data_tidanaho = Mysql(self.user, self.password, self.port, self.ip).get_tidanhao_data(
                            data_sheeps[i],
                            type_user,
                            self.oper_user, self.grade)
                        for j in range(0, len(data_tidanaho)):
                            QTreeWidgetItem(self.treewidget.topLevelItem(3).child(i)).setText(0, data_tidanaho[j])

            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(0)
            self.create_sql_tabelwidget()
            self.mainauto_match_btn.setEnabled(True)
            self.mainauto_del_btn.setEnabled(True)
            self.mainauto_add_btn.setEnabled(True)
            self.mainauto_edit_btn.setEnabled(True)
            self.mainauto_saving_btn.setEnabled(True)
            self.tableWidget.setShowGrid(True)
        else:
            self.treewidget.clear()
            self.root1 = QTreeWidgetItem(self.treewidget)
            self.root1.setText(0, '散货')
            self.root2 = QTreeWidgetItem(self.treewidget)
            self.root2.setText(0, '集装箱')
            self.root3 = QTreeWidgetItem(self.treewidget)
            self.root3.setText(0, '财务')
            self.root1.setIcon(0, user_icon)
            self.root2.setIcon(0, user_icon)
            self.root3.setIcon(0, user_icon)
            self.tableWidget.setRowCount(0)
            self.tableWidget.setColumnCount(0)
            self.tableWidget.setRowCount(2)
            self.tableWidget.setColumnCount(12)
            self.tableWidget.setSpan(0, 0, self.tableWidget.rowCount(), self.tableWidget.columnCount())
            self.tableWidget.setItem(0, 0, QTableWidgetItem('当前为离线模式  请从本地导入文件                         '))
            self.tableWidget.item(0, 0).setFont(QFont('微软雅黑', 18))
            self.tableWidget.item(0, 0).setForeground(QBrush(QColor('red')))
            self.tableWidget.item(0, 0).setTextAlignment(Qt.AlignCenter)
            self.mainauto_match_btn.setEnabled(False)
            self.mainauto_del_btn.setEnabled(False)
            self.mainauto_add_btn.setEnabled(False)
            self.mainauto_edit_btn.setEnabled(False)
            self.mainauto_saving_btn.setEnabled(False)
            self.tableWidget.item(0, 0).setFlags(Qt.NoItemFlags)
            self.pushButton.hide()
            self.pushButton_2.hide()
            self.tableWidget.horizontalHeader().hide()
            self.tableWidget.verticalHeader().hide()
            self.mainauto_result_label.setText('<font color="black" size="3">0</font>个结果')
            self.tableWidget.setShowGrid(False)

    # 初始化样式
    def mainauto_init_style(self):
        self.i = 1
        self.tableWidget.horizontalHeader().hide()
        self.tableWidget.verticalHeader().hide()
        self.pushButton.hide()
        self.pushButton_2.hide()
        # 初始化匹配结果
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)

        """
        open way
        """

    def create_sql_tabelwidget(self, mode='sql'):
        global data
        data = []
        if mode == 'sql':
            # 经理得到数据

            if int(self.grade) == 1:
                data = Mysql(self.user, self.password, self.port, self.ip).get_transport_billing_data_table()
            # 集装箱得到数据
            elif int(self.grade) == 2:
                data = Mysql(self.user, self.password, self.port, self.ip).get_sanhuo_data(self.oper_user)
            # 集装箱得到数据
            elif int(self.grade) == 3:
                data = Mysql(self.user, self.password, self.port, self.ip).get_jizhuangxiang_data(self.oper_user)
            # 财务得到数据
            elif int(self.grade) == 4:
                pass
            try:
                if len(data) == 0:
                    self.all_rows = 1


                else:
                    self.list_data = []
                    self.list_data.append('序号')
                    self.list_data.append('选择')
                    for i in data[0]:
                        self.list_data.append(i[0])
                    data1 = data
                    data = [self.list_data]
                    for i in data1[1]:
                        data.append(i)

                    self.all_rows = len(data) - 1
                    self.all_columns = len(data[0])
                    self.tableWidget.setRowCount(self.all_rows)
                    self.tableWidget.setColumnCount(self.all_columns)
                    for row in range(1, len(data)):
                        column = 2
                        for cell in data[row]:
                            self.tableWidget.setItem(row - 1, column, QTableWidgetItem(cell))
                            self.tableWidget.item(row - 1, column).setFont(QFont("微软雅黑", 10))
                            column += 1
                    for i in range(0, self.all_rows):
                        self.checkbox = QCheckBox()
                        self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                        self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                        self.serial_number_style(i)
                        self.tableWidget.setCellWidget(i, 1, self.checkbox)
                    for i in range(self.all_rows):
                        self.mytablewidget_special_item_style(i)
                    self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                    self.undo_data = []
                    self.undo_redo_count = 0

                    self.tableWidget.setHorizontalHeaderLabels(self.list_data)
                    # self.tableWidget.horizontalHeader().setStyleSheet(
                    #     "font:12pt '微软雅黑';color:rgb(59, 59, 176); background-color: rgb(85, 255, 127);")

                    # self.tableWidget.horizontalHeader().setStyleSheet(
                    #     "QHeaderView::section {background-color:lightgray;color: rgb(59, 59, 176); font: 12pt 微软雅黑; border:none;} "
                    #     "")
                    self.tableWidget.horizontalHeader().setStyleSheet(
                        "QHeaderView::section {background-color: rgb(23, 133, 175);color: rgb(255,255,255); font: 13pt 微软雅黑; border:none; border:1px dashed gray}"
                        "QHeaderView {background-color:red;}"
                    )
                    self.tableWidget.setStyleSheet("font:10pt Arial; color:gray;")
                    self.tableWidget.horizontalHeader().setMinimumHeight(43)
                    self.tableWidget.horizontalHeader().show()
                    # 用来初始化存储到tablewidget中的时间,以免多次查找下货纸,结算单等的列的序号
                    self.reseult_excel_time_colmun = {"下货纸": None, "散货结算单": None, "散货确认单": None, "产装通知": None,
                                                      "集装结算单": None, "集装确认单": None, "集港通知": None, "提单样本": None,
                                                      "装箱委托书": None}
                    try:
                        if int(self.grade) == 2:
                            for i in ["下货纸", "散货结算单", "散货确认单"]:
                                for j in range(0, self.all_columns):
                                    if self.tableWidget.horizontalHeaderItem(j).text() == i:
                                        self.reseult_excel_time_colmun[i] = j
                                        break
                        elif int(self.grade) == 3:
                            for i in ["产装通知", "集装结算单", "集装确认单", "集港通知", "提单样本", "装箱委托书"]:
                                for j in range(0, self.all_columns):
                                    if self.tableWidget.horizontalHeaderItem(j).text() == i:
                                        self.reseult_excel_time_colmun[i] = j
                                        break
                    except:
                        pass

                    # self.tableWidget.horizontalHeader().setFont(QFont('微软雅黑', 13))
                    # self.tableWidget.horizontalHeader().setForegroundRole(QColor.blue)
                    # data = Mysql(self.user, self.password, self.port, self.ip).get_transport_billing_data_table()
                    # self.all_rows = len(data) + 2
                    # self.all_columns = len(data[0]) + constant.extra_column + 1  # 这里的1是为序号留位置
                    # self.tableWidget.setRowCount(self.all_rows)
                    # self.tableWidget.setColumnCount(self.all_columns)
                    # 添加单元格数据
                    # for row in range(len(data)):
                    #     row_index = row + 2
                    #     for column in range(len(data[row])):
                    #         if column == 0:
                    #             self.checkbox = QCheckBox()
                    #             checkbox 居中显示
                    #             self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                    #             self.tableWidget.setItem(row_index, column, QTableWidgetItem(str(row + 1)))
                    #             序号的样式
                    #             self.serial_number_style(row_index)
                    #             self.tableWidget.setCellWidget(row_index, column + 1, self.checkbox)
                    #
                    #             self.tableWidget.setItem(row_index, column + 2, QTableWidgetItem(data[row][column]))
                    #         else:
                    #             self.tableWidget.setItem(row_index, column + 2, QTableWidgetItem(data[row][column]))
                # 对表头操作
                # items_list = constant.items_list
                # init_index = constant.items_list.index(constant.items_list_one_to_two[0])
                # end_index = constant.items_list.index(constant.items_list_one_to_two[-1]) + 1
                # for i in range(len(items_list)):
                #     if items_list[i] in constant.items_list_one_to_two:
                #         self.tableWidget.setItem(0, i + (i - init_index) * 2, QTableWidgetItem(items_list[i]))
                #         self.tableWidget.item(0, i + (i - init_index) * 2).setFont(QFont('微软雅黑', 13))
                #         self.tableWidget.item(0, i + (i - init_index) * 2).setForeground(QBrush(QColor('black')))
                #         self.tableWidget.setSpan(0, i + (i - init_index) * 2, 1, 3)
                #         self.tableWidget.item(0, i + (i - init_index) * 2).setTextAlignment(Qt.AlignCenter)
                #     elif i >= end_index:
                #         self.tableWidget.setItem(0, i + (end_index - init_index) * 2, QTableWidgetItem(items_list[i]))
                #         self.tableWidget.item(0, i + (end_index - init_index) * 2).setFont(QFont('微软雅黑', 13))
                #         self.tableWidget.item(0, i + (end_index - init_index) * 2).setForeground(QBrush(QColor('black')))
                #         self.tableWidget.item(0, i + (end_index - init_index) * 2).setTextAlignment(Qt.AlignCenter)
                #     else:
                #         self.tableWidget.setItem(0, i, QTableWidgetItem(items_list[i]))
                #         self.tableWidget.item(0, i).setFont(QFont("微软雅黑", 13))
                #
                #         self.tableWidget.item(0, i).setForeground(QBrush(QColor('black')))
                #         self.tableWidget.item(0, i).setTextAlignment(Qt.AlignCenter)
                #     items_list_two = constant.items_list_two
                #     for i in range(len(constant.items_list_two)):
                #         self.tableWidget.setItem(1, init_index + i, QTableWidgetItem(items_list_two[i]))
                #         self.tableWidget.item(1, init_index + i).setTextAlignment(Qt.AlignCenter)
                #         self.tableWidget.item(1, init_index + i).setFont(QFont('微软雅黑', 12))
                # self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows - 2))
                # self.undo_data = []
                # self.undo_redo_count = 0
            except:
                self.all_rows = 1

    def mytablewidget_special_item_style(self, row):
        self.tableWidget.item(row, 2).setForeground((QBrush(QColor("orange"))))
        self.tableWidget.item(row, 2).setFont(QFont('微软雅黑', 11))
        self.tableWidget.item(row, 4).setForeground(QBrush(QColor("red")))
        self.tableWidget.item(row, 4).setFont(QFont('微软雅黑', 11))
    def checkbox_index(self):
        checkbox_index = []
        for i in range(self.all_rows):
            if self.tableWidget.cellWidget(i, 1).isChecked():
                checkbox_index.append(i)
        return checkbox_index

    # 增加操作
    def mainauto_add_btn_lk(self):
        self.pushButton.show()
        self.pushButton_2.show()
        self.pushButton.setText('增一行')
        self.pushButton_2.setText('增多行')

    # 删除操作
    def mainauto_del_btn_lk(self):
        self.pushButton.show()
        self.pushButton_2.show()
        self.pushButton.setText('删一行')
        self.pushButton_2.setText('删多行')

    def mainauto_fresh_btn_lk(self):
        # replay = QMessageBox.warning(self, '警告', '刷新将会从数据库重新拿取数据,您当前的数据将会被覆盖,为防止数据丢失请先将当前的数据保存到'
        #                                          '本地，若已经保存，请忽视(点击NO即可)', QMessageBox.Ok | QMessageBox.No)
        # if replay == QMessageBox.Ok:
        #     self.save_signal_trigger()
        #
        # elif replay == QMessageBox.No:
        if self.check_db_no_conn:
            QMessageBox.critical(self, '错误', "刷新失败,原因:<strong>未连接数据库</strong>")
        elif self.check_user_no_conn:
            QMessageBox.critical(self, '错误', '刷新失败,当前<strong>未登录</strong>,请登录后再尝试')
        else:
            self.create_tree()

    # 编辑操作
    def mainauto_edit_btn_lk(self):
        self.pushButton_2.setText('退出编辑')
        self.pushButton_2.show()
        self.tableWidget.setEditTriggers(QAbstractItemView.AllEditTriggers)
        # 单元格不可编辑
        # for row in range(0, 2):
        #     for column in range(self.all_columns):
        #         try:
        #             self.tableWidget.item(row, column).setFlags(Qt.NoItemFlags)
        #         except:
        #             self.tableWidget.setItem(row, column, QTableWidgetItem(""))
        #             self.tableWidget.item(row, column).setFlags(Qt.NoItemFlags)

    def serial_number_style(self, row, column=0):
        self.tableWidget.item(row, column).setTextAlignment(Qt.AlignCenter)
        self.tableWidget.item(row, column).setFont(QFont('微软雅黑', 13))
        self.tableWidget.item(row, column).setForeground(QBrush(QColor(0, 102, 0)))

    def mainauto_function_combobox_lk(self, i):
        global root, type_user
        if i == 1:  # 打开
            self.tableWidget.disconnect()
            filename = QFileDialog.getOpenFileName(self, filter='Excel文件(*.xls)\nExcel文件(*.xlsx)\n所有文件(*)',
                                                   caption='打开文件')
            try:
                excel_data = Read().excel(path=filename[0])
                for key in excel_data:
                    data = excel_data.get(key)
                    if self.check_db_no_conn:
                        btn_check = QMessageBox.warning(self, "警告", '由于当前未连接数据库,将以默认方式导入文件,因此可能导致程序异常退出,继请选择Ok',
                                                        QMessageBox.Ok | QMessageBox.No)
                        if btn_check == QMessageBox.Ok:
                            self.tableWidget.horizontalHeader().show()
                            self.tableWidget.setRowCount(0)
                            self.tableWidget.setColumnCount(0)
                            self.all_rows = len(data) - 1
                            self.all_columns = len(data[0]) + 2
                            self.tableWidget.setRowCount(self.all_rows)
                            self.tableWidget.setColumnCount(self.all_columns)
                            self.tableWidget.setHorizontalHeaderLabels(["序号", "选择"] + data[0])
                            data = data[1:]
                            for row_index, row_data in zip(range(len(data)), data):
                                for column_index, column_data in zip(range(len(row_data)), row_data):
                                    self.tableWidget.setItem(int(row_index), int(column_index + 2),
                                                             QTableWidgetItem(str(column_data)))
                            for i in range(0, self.all_rows):
                                checkbox = QCheckBox()
                                checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                                self.tableWidget.setCellWidget(i, 1, checkbox)
                                self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                                self.serial_number_style(i)
                            for i in range(self.all_rows):
                                self.mytablewidget_special_item_style(i)
                            self.mainauto_result_label.setText(
                                '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                            self.undo_data = []
                            self.undo_redo_count = 0
                            self.mainauto_match_btn.setEnabled(True)
                            self.mainauto_del_btn.setEnabled(True)
                            self.mainauto_add_btn.setEnabled(True)
                            self.mainauto_edit_btn.setEnabled(True)
                            self.mainauto_saving_btn.setEnabled(True)
                            self.tableWidget.setShowGrid(True)
                            break


                    else:
                        column_data = Mysql(self.user, self.password, self.port, self.ip).check_coumn()
                        temp = False
                        for i in column_data:

                            if data[0] == i:
                                temp = True
                                break

                        if temp:
                            self.tableWidget.horizontalHeader().show()
                            self.tableWidget.setRowCount(0)
                            self.tableWidget.setColumnCount(0)
                            self.all_rows = len(data) - 1
                            self.all_columns = len(data[0]) + 2
                            self.tableWidget.setRowCount(self.all_rows)
                            self.tableWidget.setColumnCount(self.all_columns)
                            self.tableWidget.setHorizontalHeaderLabels(["序号", "选择"] + data[0])
                            data = data[1:]
                            for row_index, row_data in zip(range(len(data)), data):
                                for column_index, column_data in zip(range(len(row_data)), row_data):
                                    self.tableWidget.setItem(int(row_index), int(column_index + 2),
                                                             QTableWidgetItem(str(column_data)))
                            for i in range(0, self.all_rows):
                                checkbox = QCheckBox()
                                checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                                self.tableWidget.setCellWidget(i, 1, checkbox)
                                self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                                self.serial_number_style(i)

                            self.mainauto_result_label.setText(
                                '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                            self.undo_data = []
                            self.undo_redo_count = 0
                            self.mainauto_match_btn.setEnabled(True)
                            self.mainauto_del_btn.setEnabled(True)
                            self.mainauto_add_btn.setEnabled(True)
                            self.mainauto_edit_btn.setEnabled(True)
                            self.mainauto_saving_btn.setEnabled(True)
                            self.tableWidget.setShowGrid(True)
                            for i in range(self.all_rows):
                                self.mytablewidget_special_item_style(i)
                            break
                        else:
                            replay = QMessageBox.warning(self, '警告', '检测到您当前导入的Excel文件为非标准从本系统导出的文件,'
                                                                     '继续导入可能出现无法预计的情况,继续导入请选择Ok',
                                                         QMessageBox.Ok | QMessageBox.No)
                            if replay == QMessageBox.Ok:
                                try:
                                    self.tableWidget.horizontalHeader().show()
                                    self.tableWidget.setRowCount(0)
                                    self.tableWidget.setColumnCount(0)
                                    self.all_rows = len(data) - 1
                                    t = 0
                                    for i in data:
                                        if len(i) > t:
                                            t = len(i)
                                    self.all_columns = t + 2
                                    self.tableWidget.setRowCount(self.all_rows)
                                    self.tableWidget.setColumnCount(self.all_columns)
                                    self.tableWidget.setHorizontalHeaderLabels(["序号", "选择"] + data[0])
                                    data = data[1:]
                                    for row_index, row_data in zip(range(len(data)), data):
                                        for column_index, column_data in zip(range(len(row_data)), row_data):
                                            self.tableWidget.setItem(int(row_index), int(column_index + 2),
                                                                     QTableWidgetItem(str(column_data)))
                                    for i in range(0, self.all_rows):
                                        checkbox = QCheckBox()
                                        checkbox.setStyleSheet(
                                            "QCheckBox::indicator{subcontrol-position:center center;}")
                                        self.tableWidget.setCellWidget(i, 1, checkbox)
                                        self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                                        self.serial_number_style(i)

                                    self.mainauto_result_label.setText(
                                        '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                                    self.undo_data = []
                                    self.undo_redo_count = 0
                                    self.mainauto_match_btn.setEnabled(True)
                                    self.mainauto_del_btn.setEnabled(True)
                                    self.mainauto_add_btn.setEnabled(True)
                                    self.mainauto_edit_btn.setEnabled(True)
                                    self.mainauto_saving_btn.setEnabled(True)
                                    self.tableWidget.setShowGrid(True)
                                    break
                                except:
                                    pass
            except:
                pass
            self.undo_data = []
            self.undo_redo_count = 0
            self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
            self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

        elif i == 2:
            if int(self.grade) == 1:
                self.showDialog()
            else:
                QMessageBox.about(self, 'help',
                                  '<p style="line-height: 25px; font-size: 16px" ><b>添加-->增一行:</b>  会在最后一行下增加一行<br>'
                                  '<b>添加-->增多行:</b>  输入数字后会在最后一行下增加输入的行数<br>'
                                  '<b>查询-->上一个|下一个:</b>  会根据点击的结果往前后或后移动<br>'
                                  '<b>编辑-->退出编辑:</b>  当点击编辑后会进入编辑模式,点击退出编辑才会退出编辑模式<br>'
                                  '<b>编辑-->刷新:</b>  重新提取数据到当前界面<br>'
                                  '<b>删除-->删一行:</b>  默认删除最后一行<br>'
                                  '<b>删除-->删多行:</b>  会删除选中的对应行<br>'
                                  '<b>保存-->保存:</b>  会将数据上传至数据库,更新数据库内容<br>'
                                  '<b>保存-->Excel:</b>  会以表格形式保存到Excel,文件格式为.xls<br>'
                                  '<b>功能-->打开:</b>  会根据选择方式打开所选Excel文件<br>'
                                  '<b>功能-->恢复:</b>  重置所有的操作, 重置个人对节点的设置<br>'
                                  '<b>功能-->历史:</b>  打开一个新窗口的到所有历史操作完成的记录<br>'
                                  '<b>功能-->tips:</b>  得到所有的对单个提单号的操作记录(保存后的名称)'
    
                                  '<b>撤销反撤销:</b> 在不删表的情况下， undo-redo可以一直撤销或者重写，'
                                  '若果您发现数据未改变，那么可能是因为这个值是初始值(既最开始从数据中导出的数据,'
                                  '您可以使用ctrl+y快捷键实现恢复这一步骤上次更改的值，注意ctrl+z 是不会失效的，'
                                  '您每按一次就会之前走一次,没有改变是因为ctrl+z记录的值和当前值一致 '
                                  '在删表(删一行或多行的情况下)，由于数据需要，不能在更改数据新单元格后撤销到删行之前的步骤</p>')


        elif i == 3:  # 换号
            reply = QMessageBox.question(self, '提醒', "您将要执行换号操作,<strong>当前表格数据将不会被保存，</strong>继续则会先保存到本地然后执行换号",
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.save_signal_trigger()
                self.main_change_account_number_signal.emit()
            elif reply == QMessageBox.No:
                pass
        elif i == 4:  # 记录
            self.tips_signal.emit(self.grade)
        # elif i == 5:  # tips
        #     self.tips_real_signal.emit(self.grade)
        elif i == 5:  # 历史查询
            if int(self.grade) == 1:
                self.mainauto_fucntion_combobox.setCurrentIndex(0)
                self.manager_staff()
                return None
            self.history_dialog()
        elif i == 6:  # 查看同级的信息
            self.mainauto_fucntion_combobox.setCurrentIndex(0)
            self.show_same_level_data()
        elif i == 7:  # 保存到excel
            self.save_signal_trigger()
        elif i == 8:  # 退出
            self.close()
        self.mainauto_fucntion_combobox.setCurrentIndex(0)

    def closeEvent(self, event):
        reply = QMessageBox.question(self, '信息', "确认退出吗？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            # sanhuo_tips, jizhuangxiang_tips,user_sanhuo, user_jizhuangxiang
            if self.sanhuo_tips.isVisible():
                self.sanhuo_tips.hide()
            if self.jizhuangxiang_tips.isVisible():
                self.jizhuangxiang_tips.hide()
            if self.user_sanhuo.isVisible():
                self.user_sanhuo.hide()
            if self.user_jizhuangxiang.isVisible():
                self.user_jizhuangxiang.hide()

            event.accept()
        else:
            event.ignore()

    # 查看同级信息操作

    def show_same_level_data(self):
        if self.grade == 1:
            pass
        # 散货
        else:
            data = Mysql(self.user, self.password, self.port, self.ip).get_level_data(self.grade)
            print("=======================>>>>>>>>>>>>>>>>", data[0])
            self.show_same_level_data_dialog(data)

    def show_same_level_data_dialog(self, data):
        self.dialog = QDialog(self)
        self.dialog.setWindowTitle("部门信息")
        self.dialog.resize(800, 600)

        tablewidget = QTableWidget(self.dialog)
        tablewidget.setColumnCount(len(data[0]))
        tablewidget.setRowCount(len(data[1]))

        tablewidget.horizontalHeader().setStyleSheet(
            "QHeaderView::section {background-color: rgb(160, 160, 160) ;color: rgb(255,255,255); font: 13pt 黑体; border:none;border-right:0.5px dotted rgb(255,255,255);}"
        )
        for index_row, i in enumerate(data[1]):
            for index_column, j in enumerate(i):
                try:
                    tablewidget.setItem(index_row, index_column, QTableWidgetItem(j))
                except Exception as e:
                    # tablewidget.setItem(index_row, index_column, QTableWidgetItem(""))
                    print("这里出错了, 出错原因", e)
        tablewidget.setHorizontalHeaderLabels(data[0])
        layout = QVBoxLayout()
        search_btn = QPushButton("搜索")
        reload_btn = QPushButton("刷新")
        last_btn = QPushButton("上一下")
        # last_btn.hide()
        next_btn = QPushButton("上一个")
        label = QLabel("共 {} 个 结 果".format(tablewidget.rowCount()))
        search_btn.clicked.connect(lambda: self.search_same_level_data(tablewidget, label, data))
        reload_btn.clicked.connect(lambda: self.reload_same_level_data(tablewidget, label))
        next_btn.clicked.connect(lambda: self.next_same_data(tablewidget, label))
        last_btn.clicked.connect(lambda: self.last_same_data(tablewidget, label))
        # next_btn.hide()
        btn_layout = QHBoxLayout()
        layout.addWidget(tablewidget)
        layout.addLayout(btn_layout)
        self.dialog.setLayout(layout)
        btn_layout.addWidget(label)
        btn_layout.addWidget(search_btn)
        btn_layout.addWidget(reload_btn)
        btn_layout.addWidget(last_btn)
        btn_layout.addWidget(next_btn)
        self.dialog.exec_()

    # 同部门信息查询
    def search_same_level_data(self, tablewidget, label, data):
        match_text, okPressed = QInputDialog.getText(self, "内容查找", "请输入查找的内容:", QLineEdit.Normal, "")
        if match_text or okPressed:
            match_result = tablewidget.findItems(match_text, Qt.MatchExactly)
            self.search_same_level_result = match_result
            if match_result:
                tablewidget.verticalScrollBar().setSliderPosition(match_result[0].row())
                tablewidget.horizontalScrollBar().setSliderPosition(match_result[0].column())
                self.same_level_march_pos = 0
                for i in range(len(match_result)):
                    match_result[i].setBackground(QBrush(QColor('yellow')))
                match_result[0].setBackground(QBrush(QColor('lightgreen')))

                label.setText('共:<font color="red" >{}</font> 当前<font color="red" '
                              '>{}</font> '.format(len(match_result),
                                                   1))
            else:
                pass
        else:
            for index_row, i in enumerate(data[1]):
                print("----------------->i=", i)
                for index_column, j in enumerate(i):
                    try:
                        tablewidget.setItem(index_row, index_column, QTableWidgetItem(j))
                    except Exception as e:
                        print("这里出错了, 出错原因", e)

    def reload_same_level_data(self, tablewidget, label):
        data = Mysql(self.user, self.password, self.port, self.ip).get_level_data(self.grade)
        print("=======================>>>>>>>>>>>>>>>>", data[0])

        tablewidget.setColumnCount(len(data[0]))
        tablewidget.setRowCount(len(data[1]))

        tablewidget.horizontalHeader().setStyleSheet(
            "QHeaderView::section {background-color: rgb(160, 160, 160) ;color: rgb(255,255,255); font: 13pt 黑体; border:none;border-right:0.5px dotted rgb(255,255,255);}"
        )
        for index_row, i in enumerate(data[1]):
            for index_column, j in enumerate(i):
                try:
                    tablewidget.setItem(index_row, index_column, QTableWidgetItem(j))
                except Exception as e:
                    # tablewidget.setItem(index_row, index_column, QTableWidgetItem(""))
                    print("这里出错了, 出错原因", e)
        tablewidget.setHorizontalHeaderLabels(data[0])

    def next_same_data(self, tablewidget, label):
        match_result = self.search_same_level_result
        if self.same_level_march_pos == len(match_result) - 1:
            pass
        else:
            match_result[self.same_level_march_pos].setBackground(QBrush(QColor('yellow')))
            self.same_level_march_pos += 1
            match_result[self.same_level_march_pos].setBackground(QBrush(QColor('lightgreen')))
            label.setText('共:<font color="red" >{}</font> 当前<font color="red" '
                          '>{}</font> '.format(len(match_result),
                                               self.same_level_march_pos + 1))

    def last_same_data(self, tablewidget, label):
        match_result = self.search_same_level_result
        if self.same_level_march_pos == 0:
            pass
        else:
            match_result[self.same_level_march_pos].setBackground(QBrush(QColor('yellow')))
            self.same_level_march_pos -= 1
            match_result[self.same_level_march_pos].setBackground(QBrush(QColor('lightgreen')))
            label.setText('共:<font color="red" >{}</font> 当前<font color="red" '
                          '>{}</font> '.format(len(match_result),
                                               self.same_level_march_pos + 1))

    # 保存操作
    def mainauto_saving_btn_lk(self):
        """
        该函数用于保存按钮点击后的函数处理
        """
        if self.check_db_no_conn:
            QMessageBox.critical(self, '错误', "保存失败 原因:<strong> 未连接数据库</strong>")
        elif self.check_user_no_conn:
            QMessageBox.critical(self, '错误', "当前<strong>未登录</strong>,请登录后再尝试")
        else:
            self.check(quiet=False)

    def check(self, quiet: bool):
        """
        该函数用于主界面保存数据到数据库时对数据进行检测
        :param quiet: 安静模式,即保存成功不会提示
        :return: bool
        """
        list_tidanhao = []
        def check_data():
            for i in range(0, self.all_rows):
                if self.tableWidget.item(i, 2).text():
                    pass
                else:
                    reply = QMessageBox.warning(self, '警告', "<strong>船名未填</strong>, 是否定位到未填位置",
                                                QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.tableWidget.verticalScrollBar().setSliderPosition(
                            i)
                        self.tableWidget.horizontalScrollBar().setSliderPosition(
                            2)

                    return False
                if self.tableWidget.item(i, 4).text():
                    if self.tableWidget.item(i, 4).text() in list_tidanhao:
                        reply = QMessageBox.warning(self, '警告',
                                                    "<strong>提单号重复</strong>,请核对并保证所有提单号唯一, 是否定位到重复位置",
                                                    QMessageBox.Yes | QMessageBox.No)
                        if reply == QMessageBox.Yes:
                            self.tableWidget.verticalScrollBar().setSliderPosition(
                                i)
                            self.tableWidget.horizontalScrollBar().setSliderPosition(
                                4)
                        return False
                    list_tidanhao.append(self.tableWidget.item(i, 4).text())
                else:
                    reply = QMessageBox.warning(self, '警告', "<strong>提单号未填</strong>, 是否定位到未填位置",
                                                QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.tableWidget.verticalScrollBar().setSliderPosition(
                            i)
                        self.tableWidget.horizontalScrollBar().setSliderPosition(
                            4)
                    return False
                if not self.tableWidget.item(i, 3).text():
                    reply = QMessageBox.warning(self, '警告', "<strong>航次未填</strong>, 是否定位到未填位置",
                                                QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.tableWidget.verticalScrollBar().setSliderPosition(
                            i)
                        self.tableWidget.horizontalScrollBar().setSliderPosition(
                            3)
                    return False
            return True

        if check_data():
            all_data = []
            for row in range(0, self.all_rows):
                row_data = []
                for column in range(2, self.all_columns):
                    try:
                        row_data.append(self.tableWidget.item(row, column).text())
                    except:
                        row_data.append("")
                all_data.append(row_data)
            try:
                if Mysql(self.user, self.password, self.port, self.ip).save_transport_billing_data_table(
                        data=all_data, grade=self.grade, user=self.oper_user):
                    if quiet:
                        pass
                    else:
                        QMessageBox.information(self, '通知', '数据已存入数据库')
            except:
                if quiet:
                    pass
                else:
                    QMessageBox.critical(self, '错误', '由于不明原因,数据存入失败,请重试或联系管理员或QQ咨询3393370171')

    # 查找操作 match_btn_slot
    def mainauto_match_btn_lk(self):
        """断开redo 和 undo"""
        self.tableWidget.disconnect()
        """
        匹配方式
        """
        if self.match_result:
            for i in range(len(self.match_result)):
                self.match_result[i].setBackground(QBrush(QColor('white')))
                row = self.match_result[i].row()
                for column in range(2, self.all_columns):
                    self.tableWidget.item(row, column).setFont(QFont("微软雅黑", 10))
        self.pushButton.setText('上一个')
        self.pushButton_2.setText('下一个')
        self.pushButton.show()
        self.pushButton_2.show()
        match_text, okPressed = QInputDialog.getText(self, "内容查找", "请输入查找的内容:", QLineEdit.Normal, "")
        if match_text or okPressed:
            self.match_result = self.tableWidget.findItems(match_text, Qt.MatchExactly)
            if self.match_result:
                self.tableWidget.verticalScrollBar().setSliderPosition(self.match_result[0].row())
                self.tableWidget.horizontalScrollBar().setSliderPosition(self.match_result[0].column())
                self.match_result_pos = 0
                for i in range(len(self.match_result)):
                    self.match_result[i].setBackground(QBrush(QColor('yellow')))
                self.match_result[0].setBackground(QBrush(QColor('lightgreen')))
                for column in range(2, self.all_columns):
                    self.tableWidget.item(self.match_result[0].row(), column).setFont(QFont("微软雅黑", 12))

                self.mainauto_result_label.setText('共:<font color="red" size="3">{}</font> 当前<font color="red" '
                                                   'size="3">{}</font> '.format(len(self.match_result),
                                                                                1))

            else:
                self.mainauto_result_label.setText('共:<font color="red" size="3">{}</font> 当前<font color="red" '
                                                   'size="3">{}</font> '.format(0, 0))
        else:
            self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
        self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
        self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

    # function_btn_slot
    def cellenterd_lk(self, original):
        self.listWidget.hide()
        try:
            self.old_data = [original.row(), original.column(), original.text()]
        except:
            pass
            # if original.column != (0 and 1):
            #     self.old_data = [original.row(), original.column(), '']

    def mainauto_tabelwidget_lk(self, current_row, current_column):
        self.listWidget.hide()
        try:
            self.new_data = [current_row, current_column, self.tableWidget.item(current_row, current_column).text()]
            if self.new_data[0] == self.old_data[0] and self.new_data[1] == self.old_data[1]:
                self.undo_data.append([self.new_data[0], self.new_data[1], self.old_data[2], self.new_data[2]])
        except:
            pass

    # function_btn_1 slot
    def mainauto_function_btn_1_lk(self):
        global check_data
        if self.pushButton.text() == '增一行':
            self.add_one_row("", "", " ")

        elif self.pushButton.text() == '删一行':
            if self.all_rows - 1 <= 1:
                pass
            else:
                data = []
                data.append(self.all_rows - 1)
                for i in range(2, self.all_columns):
                    try:
                        data.append(self.tableWidget.item(self.all_rows - 1, i).text())
                    except:
                        data.append('')
                self.undo_data.append({'del': 1, "data": data})
                self.all_rows -= 1
                self.tableWidget.setRowCount(self.all_rows)
                self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
        elif self.pushButton.text() == '上一个':
            """断开redo 和 undo"""
            self.tableWidget.disconnect()
            if self.match_result:
                if self.match_result_pos == 0:
                    pass
                else:
                    self.match_result_pos -= 1
                    self.match_result[self.match_result_pos].setBackground(QBrush(QColor('lightgreen')))
                    self.match_result[self.match_result_pos + 1].setBackground(QBrush(QColor('yellow')))

                    self.tableWidget.verticalScrollBar().setSliderPosition(
                        self.match_result[self.match_result_pos].row())
                    self.tableWidget.horizontalScrollBar().setSliderPosition(
                        self.match_result[self.match_result_pos].column())
                    if self.match_result[self.match_result_pos].row() == self.match_result[
                        self.match_result_pos + 1].row():
                        pass
                    else:
                        for i in range(2, self.all_columns):
                            try:
                                self.tableWidget.item(self.match_result[self.match_result_pos + 1].row(), i).setFont(
                                    QFont("微软雅黑", 10))
                                self.tableWidget.item(self.match_result[self.match_result_pos].row(), i).setFont(
                                    QFont("微软雅黑", 12))
                            except:
                                pass
                self.mainauto_result_label.setText('共:<font color="red" size="3">{}</font> 当前<font color="red" '
                                                   'size="3">{}</font> '.format(len(self.match_result),
                                                                                self.match_result_pos + 1))
            self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
            self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

        # sql追加
        # elif self.pushButton.text() == 'sql追加':
        #     all_data = []
        #     for row in range(2, self.all_rows):
        #         row_data = []
        #         for column in range(2, self.all_columns):
        #             try:
        #                 row_data.append(self.tableWidget.item(row, column).text())
        #             except:
        #                 row_data.append('')
        #         all_data.append(row_data)
        #     Mysql(self.user, self.password, self.port, self.ip).save_transport_billing_data_table(all_data)

    def save_signal_trigger(self):
        global header, table_column
        try:
            filename = QFileDialog.getSaveFileName(self, filter='Excel文件(*.xls)')
            path = filename[0]
            header = []
            if self.list_data:
                header = self.list_data[2:]
            if path:
                sheet_data = []
                if header:
                    row_data = []
                    for table_column in header:
                        row_data.append(table_column)
                    sheet_data.append(row_data)
                for row in range(self.all_rows):
                    row_data = []
                    for column in range(2, self.all_columns):
                        try:
                            row_data.append(self.tableWidget.item(row, column).text())
                        except:
                            row_data.append([])
                    sheet_data.append(row_data)
                data = {
                    'sheet1': sheet_data
                }
            Write().excel(path=path, data=data)
        except:
            pass

    def mainauto_function_btn_2_lk(self):
        if self.pushButton_2.text() == '增多行':
            btn_2_add_rows, okPressed = QInputDialog.getInt(self, "增加行数", "请输入需要增加的行数:", QLineEdit.Normal, 1)
            self.all_rows += btn_2_add_rows
            if int(btn_2_add_rows):
                self.undo_data.append({'add': btn_2_add_rows})
            self.tableWidget.setRowCount(self.all_rows)
            for i in range(btn_2_add_rows):
                self.tableWidget.setItem(self.all_rows - btn_2_add_rows + i, 0,
                                         QTableWidgetItem(str(self.all_rows - btn_2_add_rows + i + 1)))
                # 增多行序号cell的样式表
                self.serial_number_style(self.all_rows - btn_2_add_rows + i)
                for j in range(2, self.all_columns):
                    self.tableWidget.setItem(self.all_rows - btn_2_add_rows + i, j, QTableWidgetItem(str('')))
                    self.tableWidget.item(self.all_rows - btn_2_add_rows + i, j).setFont(QFont("微软雅黑", 10))
            for i in range(btn_2_add_rows):
                self.checkbox = QCheckBox()
                # checkbox 居中显示
                self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                self.tableWidget.setCellWidget(self.all_rows - i - 1, 1, self.checkbox)
                self.tableWidget.setRowCount(self.all_rows)
            self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
            for i in range(btn_2_add_rows):
                self.mytablewidget_special_item_style(i + self.all_rows - btn_2_add_rows)
        elif self.pushButton_2.text() == '删多行':
            """断开redo 和 undo"""
            self.tableWidget.disconnect()
            checkbox_index = self.checkbox_index()
            data = []
            if checkbox_index:
                times = 0
                for i in checkbox_index:
                    # 因为删除导致总行数减少，因此原来的索引失效
                    data.append({"row": i})
                    for t in range(2, self.all_columns):
                        try:
                            data.append(self.tableWidget.item(i - times, t).text())
                        except:
                            data.append("")
                    self.tableWidget.removeRow(i - times)
                    times += 1
                self.all_rows -= len(checkbox_index)
                self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                # 可设置成可改，看情况
                # -------------------------------------------
                # 删除多行后重新恢复样式
                self.undo_data.append({"del_many": len(checkbox_index), "data": data})
                for i in range(self.all_rows - checkbox_index[0]):
                    self.tableWidget.setItem(checkbox_index[0] + i, 0, QTableWidgetItem(str(checkbox_index[0] + i + 1)))
                    self.serial_number_style(checkbox_index[0] + i)

            else:
                QMessageBox.warning(self, '警告', '请先勾选 <b>选择栏</b> 的按钮以删除勾选行')
            self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
            self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

        elif self.pushButton_2.text() == '退出编辑':
            self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)


        elif self.pushButton_2.text() == '下一个':
            """断开redo 和 undo"""
            self.tableWidget.disconnect()
            if self.match_result_pos < len(self.match_result) - 1:
                self.match_result_pos += 1
                self.match_result[self.match_result_pos].setBackground(QBrush(QColor('lightgreen')))
                self.match_result[self.match_result_pos - 1].setBackground(QBrush(QColor('yellow')))
                self.tableWidget.verticalScrollBar().setSliderPosition(self.match_result[self.match_result_pos].row())
                self.tableWidget.horizontalScrollBar().setSliderPosition(
                    self.match_result[self.match_result_pos].column())
                if self.match_result[self.match_result_pos].row() == self.match_result[self.match_result_pos - 1].row():
                    pass
                else:
                    for i in range(2, self.all_columns):
                        try:
                            self.tableWidget.item(self.match_result[self.match_result_pos - 1].row(), i).setFont(
                                QFont("微软雅黑", 10))
                            self.tableWidget.item(self.match_result[self.match_result_pos].row(), i).setFont(
                                QFont("微软雅黑", 12))
                        except:
                            pass

                self.mainauto_result_label.setText('共:<font color="red" size="3">{}</font> 当前<font color="red" '
                                                   'size="3">{}</font> '.format(len(self.match_result),
                                                                                self.match_result_pos + 1))
            self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
            self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

            # btn_2_del_rows, okPressed = QInputDialog.getInt(self, "删除行数", "请输入需要删除的行数:", QLineEdit.Normal, 1)
            # self.all_rows -= btn_2_del_rows
            # self.tableWidget.setRowCount(self.all_rows)

        # sql重写
        # elif self.pushButton_2.text() == 'sql重写':
        #     all_data = []
        #     for row in range(2, self.all_rows):
        #         row_data = []
        #         for column in range(2, self.all_columns):
        #             row_data.append(self.tableWidget.item(row, column).text())
        #         all_data.append(row_data)
        #     Mysql(self.user, self.password, self.port, self.ip).save_rewrite_transport_billing_data_table(all_data)

    # 键盘控制事件
    def keyPressEvent(self, event):
        if event.key() == (Qt.Key_Control and Qt.Key_Z):
            self.undo()
        elif event.key() == (Qt.Key_Control and Qt.Key_Y):
            self.redo()
        elif event.key() == (Qt.Key_Control and Qt.Key_I):  # 增加行
            self.mainauto_add_btn_lk()
        elif event.key() == (Qt.Key_Control and Qt.Key_S):
            self.mainauto_saving_btn_lk()
        elif event.key() == (Qt.Key_Control and Qt.Key_D):
            self.mainauto_del_btn_lk()
        elif event.key() == (Qt.Key_Control and Qt.Key_N):
            self.mainauto_function_combobox_lk(1)
        elif event.key() == (Qt.Key_Control and Qt.Key_O):
            self.mainauto_function_combobox_lk(2)
        elif event.key() == (Qt.Key_Control and Qt.Key_H):
            self.mainauto_function_combobox_lk(3)
        elif event.key() == (Qt.Key_Control and Qt.Key_Q):
            self.mainauto_function_combobox_lk(4)
        elif event.key() == (Qt.Key_Control and Qt.Key_C):
            items = self.tableWidget.selectedItems()
            if math.modf(int(len(items)) / int(self.all_rows))[0] != 0.0 and len(items):
                QMessageBox.warning(self, "警告", "请仅选中列以保存")
            else:
                global header, table_column
                try:
                    filename = QFileDialog.getSaveFileName(self, filter='Excel文件(*.xls)')
                    path = filename[0]
                    if path:
                        Workbook = openpyxl.Workbook()
                        Worksheet = Workbook.active

                        for index, i in zip(range(0, len(items)), items):
                            try:
                                if i.text():
                                    t = i.text()
                                    pass
                                else:
                                    t = ""
                            except:
                                t = ""
                            Worksheet.cell((index % self.all_rows) + 1, int(index / self.all_rows) + 1, t)
                        Worksheet.insert_rows(0, 1)
                        for i in range(0, int(len(items) / self.all_rows)):
                            Worksheet.cell(1, i + 1, self.list_data[items[i * self.all_rows].column()])
                        Workbook.save(path)
                except:
                    pass
        else:
            pass

    def undo(self):
        global row, column
        self.tableWidget.disconnect()

        # 用来检测self.undo_data 是否改变，因为这里不能直接写,改变了清除原有数据直接让self.undo_data 只保存最后一次的操作
        if self.old_all_undo_stept != len(self.undo_data):
            self.undo_redo_count = 0
        self.old_all_undo_stept = len(self.undo_data)
        self.undo_redo_count -= 1
        if self.undo_data:
            try:
                data = self.undo_data[self.undo_redo_count]
                if type(data) == list:
                    try:
                        self.tableWidget.item(data[0], data[1]).setText('{}'.format(data[2]))
                    except:
                        pass
                else:
                    key_list = []
                    for key in data:
                        key_list.append(key)
                    if 'add' in key_list:
                        self.all_rows -= data.get('add')
                        self.tableWidget.setRowCount(self.all_rows)
                        self.mainauto_result_label.setText(
                            '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))

                    elif 'del' in key_list:
                        if data.get('del') == 1:
                            try:
                                if self.tableWidget.item(data.get('data')[0], 0).text():
                                    undo_data_end = self.undo_data[-1]
                                    self.undo_data = []
                                    self.undo_redo_count = -1
                                    self.undo_data.append(undo_data_end)


                            except:
                                self.all_rows += 1
                                self.tableWidget.setRowCount(self.all_rows)
                                for i in range(1, len(data.get('data'))):
                                    self.tableWidget.setItem(self.all_rows - 1, i + 1,
                                                             QTableWidgetItem('{}'.format(data.get('data')[i])))
                                self.checkbox = QCheckBox()
                                # checkbox 居中显示
                                self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                                self.tableWidget.setCellWidget(self.all_rows - 1, 1, self.checkbox)
                                self.tableWidget.setItem(self.all_rows - 1, 0,
                                                         QTableWidgetItem('{}'.format(data.get('data')[0] + 1)))
                                self.serial_number_style(self.all_rows - 1)
                                self.mainauto_result_label.setText(
                                    '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                                for i in range(self.all_rows):
                                    self.mytablewidget_special_item_style(i)

                    elif 'del_many' in key_list:
                        check = []
                        for i in range(0, self.all_rows):
                            try:
                                check.append(self.tableWidget.item(i, 4).text())
                            except:
                                pass
                        if data.get('data')[3] in check:
                            pass
                        else:
                            for i in range(len(data.get('data'))):
                                if type(data.get('data')[i]) == dict:
                                    row = data.get('data')[i].get('row')
                                    column = 2
                                    self.tableWidget.insertRow(row)
                                    self.all_rows += 1
                                    self.tableWidget.setRowCount(self.all_rows)
                                    # 样式设置
                                    checkbox = QCheckBox()
                                    checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                                    self.tableWidget.setCellWidget(row, 1, checkbox)
                                else:
                                    self.tableWidget.setItem(row, column, QTableWidgetItem(str(data.get('data')[i])))
                                    column += 1
                            for row in range(0, self.all_rows):
                                self.tableWidget.setItem(row, 0, QTableWidgetItem(str(row + 1)))
                                self.serial_number_style(row)
                            self.mainauto_result_label.setText(
                                '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                            for i in range(self.all_rows):
                                self.mytablewidget_special_item_style(i)



            except:
                self.undo_redo_count += 1
                pass
        self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
        self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

    def redo(self):
        global column, row
        self.tableWidget.disconnect()
        if self.old_all_undo_stept != len(self.undo_data):
            self.undo_redo_count = 0
        self.old_all_undo_stept = len(self.undo_data)
        if self.undo_redo_count >= 0:
            self.undo_redo_count = 0
        elif self.undo_data:
            try:
                data = self.undo_data[self.undo_redo_count]
                if type(data) == list:
                    self.tableWidget.item(data[0], data[1]).setText('{}'.format(data[3]))
                else:
                    key_list = []
                    for key in data:
                        key_list.append(key)
                    if 'add' in key_list:
                        for i in range(data.get('add')):
                            self.all_rows += 1
                            self.tableWidget.setRowCount(self.all_rows)
                            self.checkbox = QCheckBox()
                            self.tableWidget.setCellWidget(self.all_rows - 1, 1, self.checkbox)
                            self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
                            self.tableWidget.setItem(self.all_rows - 1, 0, QTableWidgetItem(str(self.all_rows)))
                            self.serial_number_style(self.all_rows - 1)
                            for j in range(2, self.all_columns):
                                self.tableWidget.setItem(self.all_rows - 1, j, QTableWidgetItem(str('')))
                            for k in range(self.all_rows):
                                self.mytablewidget_special_item_style(k)
                    elif 'del' in key_list:
                        self.all_rows -= 1
                        self.tableWidget.setRowCount(self.all_rows)
                        self.mainauto_result_label.setText(
                            '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
                    elif 'del_many' in key_list:
                        for i in range(len(data.get('data'))):
                            if type(data.get('data')[i]) == dict:
                                row = data.get('data')[i].get('row')
                                self.tableWidget.removeRow(row)
                                self.all_rows -= 1
                                self.tableWidget.setRowCount(self.all_rows)
                        for row in range(2, self.all_rows):
                            self.tableWidget.setItem(row, 0, QTableWidgetItem(str(row + 1)))
                            self.serial_number_style(row)
                    self.mainauto_result_label.setText(
                        '<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
            except:
                pass
            self.undo_redo_count += 1
        self.tableWidget.currentItemChanged.connect(self.cellenterd_lk)
        self.tableWidget.cellChanged.connect(self.mainauto_tabelwidget_lk)

    # 以下是对当管理员添加新用户时的处理

    def showDialog(self):
        self.dialog = QDialog()
        self.dialog.setWindowTitle("添加")
        nameLabel = QLabel(" &用户", self.dialog)
        nameLineEdit = QLineEdit(self.dialog)
        passwordLabel = QLabel(" &密码", self.dialog)
        passwordLineEdit = QLineEdit(self.dialog)

        gradeLabel = QLabel(" &权限", self.dialog)
        spainBox = QSpinBox(self.dialog)
        spainBox.setMinimum(1)
        spainBox.setMaximum(4)
        # 设置伙伴控件
        nameLabel.setBuddy(nameLineEdit)
        passwordLabel.setBuddy(passwordLineEdit)
        gradeLabel.setBuddy(spainBox)

        buttonOk = QPushButton('确定', self.dialog)
        buttonCancle = QPushButton("重置", self.dialog)
        mainLayout = QGridLayout(self.dialog)
        mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addWidget(nameLineEdit, 0, 1, 1, 2)
        mainLayout.addWidget(passwordLabel, 1, 0)
        mainLayout.addWidget(passwordLineEdit, 1, 1, 1, 2)
        mainLayout.addWidget(gradeLabel, 2, 0)
        mainLayout.addWidget(spainBox, 2, 1, 1, 2)
        mainLayout.addWidget(buttonOk, 3, 1)
        mainLayout.addWidget(buttonCancle, 3, 2)
        buttonOk.clicked.connect(lambda: self.OK_pressed(nameLineEdit, passwordLineEdit, spainBox))
        buttonCancle.clicked.connect(lambda: self.Cancle_pressed(nameLineEdit, passwordLineEdit))
        self.dialog.exec_()

    # 以下是管理员对员工账号的管理,更名,修改密码,调整等级
    def manager_staff(self):
        self.dialog = QDialog()
        self.dialog.setWindowTitle("管理")
        nameLabel = QLabel(" &旧用户", self.dialog)
        nameLineEdit = QLineEdit(self.dialog)
        passwordLabel = QLabel(" &旧密码", self.dialog)
        passwordLineEdit = QLineEdit(self.dialog)

        newnameLabel = QLabel(" &新用户", self.dialog)
        newnameLineEdit = QLineEdit(self.dialog)
        newpasswordLabel = QLabel(" &新密码", self.dialog)
        newpasswordLineEdit = QLineEdit(self.dialog)

        # 设置伙伴控件
        nameLabel.setBuddy(nameLineEdit)
        passwordLabel.setBuddy(passwordLineEdit)

        newnameLabel.setBuddy(newnameLineEdit)
        newpasswordLabel.setBuddy(newpasswordLineEdit)


        buttonOk = QPushButton('确定', self.dialog)
        buttonCancle = QPushButton("重置", self.dialog)
        mainLayout = QGridLayout(self.dialog)
        mainLayout.addWidget(nameLabel, 0, 0)
        mainLayout.addWidget(nameLineEdit, 0, 1, 1, 2)

        mainLayout.addWidget(newnameLabel, 0, 4)
        mainLayout.addWidget(newnameLineEdit, 0, 5, 1, 2)

        mainLayout.addWidget(passwordLabel, 1, 0)
        mainLayout.addWidget(passwordLineEdit, 1, 1, 1, 2)

        mainLayout.addWidget(newpasswordLabel, 1, 4)
        mainLayout.addWidget(newpasswordLineEdit, 1, 5, 1, 2)

        mainLayout.addWidget(buttonOk, 2, 2)
        mainLayout.addWidget(buttonCancle, 2, 5)
        buttonOk.clicked.connect(lambda: self.OK_pressed(nameLineEdit, passwordLineEdit, newuser=newnameLineEdit, newpassword=newpasswordLineEdit, spinbox="", check=True))
        buttonCancle.clicked.connect(lambda: self.Cancle_pressed(nameLineEdit, passwordLineEdit, newnameLineEdit, newpasswordLineEdit))
        self.dialog.exec_()


    def OK_pressed(self, user, password, spinbox, newuser=False, newpassword=False, check=False):
        if not check:
            if user.text() == "" or password.text() == "":
                QMessageBox.critical(self.dialog, "错误", "用户名和密码均不能为空")
            else:
                list_user = Mysql(self.user, self.password, self.port, self.ip).get_user_password_data(user=True)
                for i in list_user:
                    re_user = "^" + user.text() + "$"
                    if re.findall(re_user, i, re.I):
                        QMessageBox.critical(self.dialog, "错误", "该用户已存在")
                        return 0
                res = Mysql(self.user, self.password, self.port, self.ip).add_user(user.text(), password.text(),
                                                                                   spinbox.text())
                if int(spinbox.text()) == 1:
                    pass
                elif int(spinbox.text()) == 2:
                    pass
                elif int(spinbox.text()) == 3:
                    pass
                if res:
                    QMessageBox.information(self.dialog, "information", "添加成功")
        else:
            if user.text()=="" or password.text()=="" or newuser.text()=="" or newpassword.text()=="":
                QMessageBox.critical(self.dialog, "错误", "用户名和密码均不能为空")
            else:
                # list_user = Mysql(self.user, self.password, self.port, self.ip).get_user_password_data(user=True)
                # for i in list_user:
                #     re_user = "^" + user.text() + "$"
                #     if not re.findall(re_user, i, re.I):
                #
                #         return 0
                res = Mysql(self.user, self.password, self.port, self.ip).manager_staff(user.text(), password.text(),
                                                                                   newuser.text(), newpassword.text(),)
                if res:
                    QMessageBox.information(self.dialog, "information", "修改成功")
                else:
                    QMessageBox.critical(self.dialog, "错误", "该用户不存在")
    # 管理员添加新用户时的选择重置按钮
    def Cancle_pressed(self, user, password, newuser=False, newpassword=False):
        user.setText("")
        password.setText("")
        if newpassword or newuser:
            newpassword.setText("")
            newuser.setText("")

    def insert_time_date(self, shipname, tidanhao, new_tidanhao, combobox, data):
        if tidanhao != new_tidanhao:
            try:
                tidanhao_row = self.tableWidget.findItems(new_tidanhao, Qt.MatchExactly)[0].row()
            except:
                return None
            if tidanhao_row:
                tidanhao = new_tidanhao
            else:
                return None

        tidanhao_row = self.tableWidget.findItems(tidanhao, Qt.MatchExactly)[0].row()
        for i in combobox:
            temp = i  # 因为combobox不能作为column直接提取作为dict的key,所以要做一个临时变量
            if i == "确认单" or i == "结算单" or i == "费用确认单" or i == "费用结算单":
                if int(self.grade) == 2:
                    i = "散货" + i
                elif int(self.grade) == 3:
                    i = "集装" + i[2:]
            if data.get(temp):
                try:
                    self.tableWidget.item(tidanhao_row, self.reseult_excel_time_colmun.get(i)).setText(data.get("time"))
                except:
                    self.tableWidget.setItem(tidanhao_row, self.reseult_excel_time_colmun.get(i),
                                             QTableWidgetItem(data.get("time")))

    def history_dialog(self):
        self.history_win_dialog = QDialog()
        self.history_win_dialog.resize(800, 600)
        self.history_win_dialog.setWindowTitle("历史查询")
        self.history_win_dialog.setWindowIcon(QIcon(":/open.ico"))
        self.history_match_btn = QPushButton()
        self.history_match_btn.setText("查询")
        self.history_export_btn = QPushButton()
        self.history_export_btn.setText("导出")
        self.history_lable = QLabel()
        self.history_last_btn = QPushButton("上一个")
        self.history_next_btn = QPushButton("下一个")
        self.history_label = QLabel()
        self.history_last_btn.clicked.connect(lambda: self.history_last_lk(history_tablewidget))
        self.history_next_btn.clicked.connect(lambda: self.history_next_lk(history_tablewidget))
        self.history_match_btn.clicked.connect(lambda: self.history_match_lk(history_tablewidget))
        self.history_export_btn.clicked.connect(lambda: self.history_export_lk(history_tablewidget))
        history_tablewidget = QTableWidget()
        data = Mysql(self.user, self.password, self.port, self.ip).history_find(self.oper_user, self.grade)
        history_tablewidget.setColumnCount(len(data[0]))
        history_tablewidget.setRowCount(len(data[1]))
        data_data = data[1]
        history_tablewidget.setHorizontalHeaderLabels(data[0])
        for row_index, data_row_data in zip(range(0, len(data_data)), data_data):
            for column_index, data_column_data in zip(range(0, len(data_row_data)), data_row_data):
                history_tablewidget.setItem(row_index, column_index, QTableWidgetItem(data_column_data))
                history_tablewidget.item(row_index, column_index).setFont(QFont("微软雅黑", 10))
        # history_tablewidget.verticalHeader().hide()
        layout_btn = QHBoxLayout()
        layout_btn.addWidget(self.history_label)
        layout_btn.addWidget(self.history_match_btn)
        layout_btn.addWidget(self.history_export_btn)
        layout_btn.addWidget(self.history_next_btn)
        layout_btn.addWidget(self.history_last_btn)
        self.history_label.setText(" 共 {} 个 结 果".format(history_tablewidget.rowCount()))
        history_tablewidget.horizontalHeader().setStyleSheet(
            "QHeaderView::section {background-color: rgb(160, 160, 160) ;color: rgb(255,255,255); font: 13pt 黑体; border:none;border-right:0.5px dotted rgb(255,255,255);}"
        )
        self.history_next_btn.hide()
        self.history_last_btn.hide()

        layout_table = QHBoxLayout()
        layout_table.addWidget(history_tablewidget)
        centerlayout = QVBoxLayout(self.history_win_dialog)

        centerlayout.addLayout(layout_table)
        centerlayout.addLayout(layout_btn)
        history_tablewidget.horizontalHeader().setFont(QFont('微软雅黑', 13))
        self.history_win_dialog.setLayout(centerlayout)
        self.history_win_dialog.exec_()

    def history_match_lk(self, history_tablewidegt):
        self.history_last_btn.show()
        self.history_next_btn.show()
        match_text, okPressed = QInputDialog.getText(self.history_win_dialog, "内容查找", "请输入查找的内容:", QLineEdit.Normal, "")
        if match_text or okPressed:
            try:
                for i in self.history_find_list:
                    i.setBackground(QBrush(QColor('white')))
            except:
                pass
        else:
            try:
                for i in self.history_find_list:
                    i.setBackground(QBrush(QColor('white')))
                return None
            except:
                pass
        self.history_find_list = history_tablewidegt.findItems(match_text, Qt.MatchExactly)
        if self.history_find_list:
            history_tablewidegt.verticalScrollBar().setSliderPosition(self.history_find_list[0].row())
            history_tablewidegt.horizontalScrollBar().setSliderPosition(self.history_find_list[0].column())
            self.history_find_result_pos = 0
            for i in range(len(self.history_find_list)):
                self.history_find_list[i].setBackground(QBrush(QColor('yellow')))
            self.history_find_list[0].setBackground(QBrush(QColor('lightgreen')))
            for column in range(0, history_tablewidegt.columnCount()):
                history_tablewidegt.item(self.history_find_list[0].row(), column).setFont(QFont("微软雅黑", 12))
        if self.history_find_list:
            self.history_label.setText(" 共 {}  ".format(len(self.history_find_list)) + " 当 前 1")
        else:
            self.history_label.setText(" 共 {}  ".format(len(self.history_find_list)) + " 当  前 0")

    def history_next_lk(self, history_tablewidegt):  # 历史记录查找下一个
        if self.history_find_result_pos == len(self.history_find_list) - 1:
            pass
        else:
            self.history_find_result_pos += 1
            self.history_find_list[self.history_find_result_pos].setBackground(QBrush(QColor('lightgreen')))
            self.history_find_list[self.history_find_result_pos - 1].setBackground(QBrush(QColor('yellow')))
            history_tablewidegt.verticalScrollBar().setSliderPosition(
                self.history_find_list[self.history_find_result_pos].row())
            history_tablewidegt.horizontalScrollBar().setSliderPosition(
                self.history_find_list[self.history_find_result_pos].column())
            if self.history_find_list[self.history_find_result_pos].row() == self.history_find_list[
                self.history_find_result_pos - 1].row():
                pass
            else:
                for i in range(0, history_tablewidegt.columnCount()):
                    try:
                        history_tablewidegt.item(self.history_find_list[self.history_find_result_pos - 1].row(),
                                                 i).setFont(
                            QFont("微软雅黑", 10))
                        history_tablewidegt.item(self.history_find_list[self.history_find_result_pos].row(), i).setFont(
                            QFont("微软雅黑", 12))
                    except:
                        pass
            self.history_label.setText(
                " 共 {}  ".format(len(self.history_find_list)) + " 当 前 {}".format(self.history_find_result_pos + 1))

    def history_last_lk(self, history_tablewidget):  # 历史记录查找上一个
        if self.history_find_list:
            if self.history_find_result_pos == 0:
                pass
            else:
                self.history_find_result_pos -= 1
                self.history_find_list[self.history_find_result_pos].setBackground(QBrush(QColor('lightgreen')))
                self.history_find_list[self.history_find_result_pos + 1].setBackground(QBrush(QColor('yellow')))

                history_tablewidget.verticalScrollBar().setSliderPosition(
                    self.history_find_list[self.history_find_result_pos].row())
                history_tablewidget.horizontalScrollBar().setSliderPosition(
                    self.history_find_list[self.history_find_result_pos].column())
                if self.history_find_list[self.history_find_result_pos].row() == self.history_find_list[
                    self.history_find_result_pos + 1].row():
                    pass
                else:
                    for i in range(0, history_tablewidget.columnCount() - 2):
                        try:
                            history_tablewidget.item(self.history_find_list[self.history_find_result_pos + 1].row(),
                                                     i).setFont(
                                QFont("微软雅黑", 10))
                            history_tablewidget.item(self.history_find_list[self.history_find_result_pos].row(),
                                                     i).setFont(
                                QFont("微软雅黑", 12))
                        except:
                            pass
                self.history_label.setText(
                    " 共 {}  ".format(len(self.history_find_list)) + " 当 前 {}".format(self.history_find_result_pos + 1))

    def history_export_lk(self, history_tablewidget):
        global header, table_column
        try:
            filename = QFileDialog.getSaveFileName(self, filter='Excel文件(*.xls)')
            path = filename[0]
            header = []
            if self.list_data:
                header = self.list_data[2:]
            if path:
                sheet_data = []
                if header:
                    row_data = []
                    for table_column in header:
                        row_data.append(table_column)
                    sheet_data.append(row_data)
                for row in range(self.all_rows):
                    row_data = []
                    for column in range(0, self.all_columns - 2):
                        try:
                            row_data.append(history_tablewidget.item(row, column).text())
                        except:
                            row_data.append([])
                    sheet_data.append(row_data)
                data = {
                    'sheet1': sheet_data
                }
            Write().excel(path=path, data=data)
        except:
            pass

    # 增一行,由于在listwidget有用到,所以提出来做成函数

    def add_one_row(self, ship: str, ship_num: str, tidanhao):
        self.undo_data.append({'add': 1})
        self.all_rows += 1
        # 新增行添加QCheckBox
        self.tableWidget.setRowCount(self.all_rows)
        self.checkbox = QCheckBox()
        # checkbox 居中显示
        self.checkbox.setStyleSheet("QCheckBox::indicator{subcontrol-position:center center;}")
        self.tableWidget.setCellWidget(self.all_rows - 1, 1, self.checkbox)
        self.mainauto_result_label.setText('<font color="red" size="3">{}</font>个结果'.format(self.all_rows))
        self.tableWidget.setItem(self.all_rows - 1, 0, QTableWidgetItem(str(self.all_rows)))
        # 新增行序号的样式
        self.serial_number_style(self.all_rows - 1)
        # 为空的单元格添加""数据
        for i in range(2, self.all_columns):
            self.tableWidget.setItem(self.all_rows - 1, i, QTableWidgetItem(str('')))
            self.tableWidget.item(self.all_rows - 1, i).setFont(QFont("微软雅黑", 10))
        if ship and ship_num and tidanhao:
            self.tableWidget.setItem(self.all_rows - 1, 2, QTableWidgetItem(str(ship)))
            self.tableWidget.setItem(self.all_rows - 1, 3, QTableWidgetItem(str(ship_num)))
            self.tableWidget.setItem(self.all_rows - 1, 4, QTableWidgetItem(str(tidanhao)))
            self.check(True)
            self.create_tree()
        self.mytablewidget_special_item_style(self.all_rows - 1)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = Main_auto("root", "root", 3306, "127.0.0.1", user_pw_table_User="111",
                     user_pw_table_Pw=2222)
    main.show()
    app.exit(app.exec_())
