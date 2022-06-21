import re
import sys
import openpyxl
from PyQt5.QtGui import QWheelEvent, QIcon
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from resource.ui.usr_operation import Ui_Form
from mainauto_read_wite_file import form_operation
import os
from mysql import Mysql
import datetime
import myopne_icon_rc

"""
    问题出在第二个，和第一个在一起时无法识别
"""


class User_operation(Ui_Form, QWidget):
    textedit_finished = pyqtSignal()
    remember_history = pyqtSignal(str, int, str, str, str, list, dict)

    def __init__(self, g_User, g_Pw, g_Port, g_Ip, username, password, parent=None, *args, **kwargs):
        super(User_operation, self).__init__(parent)
        self.setWindowIcon(QIcon(":/open.ico"))
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setupUi(self)
        self.setWindowTitle("物流管理系统")
        self.history_data = {}
        self.g_User = g_User
        self.g_Pw = g_Pw
        self.g_Port = g_Port
        self.g_Ip = g_Ip
        self.user = username
        self.password = password
        self.object_name_list = []
        self.listwidget.hide()
        self.sheepname_lineedit.setFocus()
        self.operation_feiyong_combobox.currentIndexChanged.connect(self.feiyongdan_form)
        t = 1
        self.combobox_data = ["下货纸"]
        try:
            while True:
                self.combobox_data.append(self.operation_feiyong_combobox.itemText(t).strip())
                if not self.operation_from_combobox.itemText(t):
                    break
                t += 1
        except:
            pass
        self.combobox_data = [i for i in self.combobox_data if i != ""]
        self.history_data = {}

    def form_lk(self, index):
        rename, okPressed = QInputDialog.getText(self, self.operation_from_combobox.itemText(index), "请输入保存后的文件名称:",
                                                 QLineEdit.Normal, "")
        if int(index) != 6:
            data = [self.tidanhao_lineedit.text(), self.sheepname_lineedit.text(), self.chuanci_lineedit.text(),
                    self.mudigang_lineedit.text(), self.fahuoren_textedit.toPlainText(),
                    self.shouhuoren_textedit.toPlainText(), self.tongzhiren_textedit.toPlainText(),
                    self.maitou_textedit.toPlainText(), self.jianshu_lineedit.text(), self.zhongliang_lineedit.text(),
                    self.huoming_lineedit.text(), self.tiji_lineedit.text()]
            name = self.operation_from_combobox.itemText(index).strip()
            file_name = name + "下货纸.xlsx"

            if okPressed or rename:
                self.history_data["下货纸"] = rename
                form_operation().copy_template(template=file_name, data=data, rename=rename + ".xlsx")
            self.operation_from_combobox.disconnect()
            self.operation_from_combobox.setCurrentIndex(0)
            self.operation_from_combobox.currentIndexChanged.connect(self.form_lk)
        elif int(index) == 6:
            data_box = [self.tidanhao_lineedit.text(), self.fahuoren_textedit.toPlainText(),
                        self.shouhuoren_textedit.toPlainText(), self.tongzhiren_textedit.toPlainText(),
                        self.jianshu_lineedit.text(), self.zhongliang_lineedit.text(),
                        "品  名：钢 管\n现  场：张师傅 156 2302 2698\n绑  扎：白师傅13759682365\n刘师傅15152023316"
                        ]
            data_pra = [self.sheepname_lineedit.text(), self.chuanci_lineedit.text(),
                        self.mudigang_lineedit.text(), self.maitou_textedit.toPlainText(), self.huoming_lineedit.text()]
            # rename, okPressed = QInputDialog.getText(self, self.operation_from_combobox.itemText(index), "请输入保存后的文件名称:",
            #                                          QLineEdit.Normal, "")
            if okPressed or rename:
                self.history_data["下货纸"] = rename
                form_operation().sanhuo_docx_file("外运", rename, data_pra, data_box)
            self.operation_from_combobox.disconnect()
            self.operation_from_combobox.setCurrentIndex(0)
            self.operation_from_combobox.currentIndexChanged.connect(self.form_lk)

    def feiyongdan_form(self, index):
        global save_name, okPressed, filename

        if index != 0:
            filename = self.operation_feiyong_combobox.itemText(index).strip()
            save_name, okPressed = QInputDialog.getText(self, filename, "请输入保存后的文件名称:", QLineEdit.Normal, "")
        if save_name or okPressed:

            data = []
            # 结算单
            if index == 1:
                data.append("装船日：{}".format(datetime.date.today().strftime("%Y/%m/%d")))
                data.append(self.kehumingcheng_lineedit.text())
                data.append(self.sheepname_lineedit.text() + "/" + self.chuanci_lineedit.text())
                data.append(self.tidanhao_lineedit.text())
                data.append(self.qiyungang_lineedit.text())
                data.append(self.mudigang_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.tiji_lineedit.text())
                data.append(self.jifeidun_lineedit.text())
                # 海运费
                data.append(self.jifeidun_lineedit.text())
                data.append(self.jifeidun_lineedit.text())
                data.append(self.chuandongmingcheng_lineedit.text())
                # 港杂费
                data.append(self.zhongliang_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.dimiangongsi_lineedit.text())
                # 绑扎费
                data.append(self.jifeidun_lineedit.text())
                data.append(self.jifeidun_lineedit.text())
                data.append(self.bangzagongsi_lineedit.text())
                # 文件费
                data.append(self.chuandaigongsi_lineedit.text())
                # 苫盖费
                data.append(self.zhongliang_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.bangzagongsi_lineedit.text())
                # 电放费
                data.append(self.chuandaigongsi_lineedit.text())
                # 改单费
                data.append(self.chuandaigongsi_lineedit.text())
                # 过磅费
                data.append(self.zhongliang_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.dimiangongsi_lineedit.text())
                # 堆存费
                data.append(self.zhongliang_lineedit.text())
                data.append(self.zhongliang_lineedit.text())
                data.append(self.dimiangongsi_lineedit.text())
                # 开票日期
                data.append(datetime.date.today().strftime("%Y/%m/%d"))
                self.history_data["结算单"] = save_name
            # 确认单
            if index == 2:
                data.append(self.tidanhao_lineedit.text())
                data.append(self.sheepname_lineedit.text() + "/" + self.chuanci_lineedit.text())
                data.append(self.qiyungang_lineedit.text())
                data.append(self.mudigang_lineedit.text())
                data.append(datetime.date.today().strftime("%Y/%m/%d"))
                data.append(self.zhongliang_lineedit.text())

                # 海运费
                data.append(self.jifeidun_lineedit.text())
                # 港杂费
                data.append(self.zhongliang_lineedit.text())
                # 绑扎费
                data.append(self.jifeidun_lineedit.text())
                # 苫盖费
                data.append(self.zhongliang_lineedit.text())
                # 过磅费
                data.append(self.zhongliang_lineedit.text())
                # 堆存费
                data.append(self.zhongliang_lineedit.text())

                data.append(datetime.date.today().strftime("%Y/%m/%d"))
                self.history_data["确认单"] = save_name

            if index > 0:
                form_operation().copy_feiyongdan_file(filename, save_name, data)
                self.operation_feiyong_combobox.setCurrentIndex(0)
        else:
            # 需要重置index
            self.operation_feiyong_combobox.setCurrentIndex(0)
            pass

    def refresh_lk(self):
        data = Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).sanhuo_data(tidanhao=self.text)
        data = data[0]
        self.create_nwe(self.text, data)

    def create_nwe(self, text, data):
        self.reset_lk()
        """
            船名
            航次
            提单号
            IMO号
            启运港
            目的港
            起运港ETA
            起运港ATB
            起运港ATD
            目的港ETA
            
            货物类型
            客户名称
            
            船代公司
            联系人
            报关公司
            地面公司
            特殊要求
            绑扎公司
            特殊要求
            船东名称
            
            发货状态
            出单方式
            海运条款
            
            货名
            件数
            重量
            体积
            计费吨
            发货人
            收货人
            通知人
            唛头
            货描
        """

        self.ship = data[0]
        self.text = text
        data1 = data[:17]
        data2 = data[20:]
        data_button = data[17:20]
        data = data1 + data2
        self.list1 = [self.sheepname_lineedit, self.chuanci_lineedit, self.tidanhao_lineedit,
                      self.IMOhao_lineedit, self.qiyungang_lineedit, self.mudigang_lineedit,
                      self.qiyungangETA_lineedit, self.qiyungangATB_lineedit,
                      self.qiyungangATD_lineedit, self.mudigangETA_lineedit,
                      self.huwuleixing_lineedit, self.kehumingcheng_lineedit,
                      self.chuandaigongsi_lineedit,
                      self.baoguangongsi_lineedit, self.dimiangongsi_lineedit,
                      self.bangzagongsi_lineedit, self.chuandongmingcheng_lineedit,
                      self.huoming_lineedit, self.jianshu_lineedit, self.zhongliang_lineedit,
                      self.tiji_lineedit, self.jifeidun_lineedit, self.fahuoren_textedit,
                      self.shouhuoren_textedit, self.tongzhiren_textedit, self.maitou_textedit,
                      self.huomiao_textedit]

        # 对单选框和多选框的处理
        btn_check = [self.operation_fahuo_checkbox, self.operation_daohuo_checkbox,
                     self.operation_dizai_checkbox, self.checkBox_18, self.operation_jiancha_checkbox,
                     self.operation_fangxing_checkbox]

        for i in range(len(btn_check)):
            if btn_check[i].text() == data_button[0]:
                for btn in range(len(btn_check[:i + 1])):
                    btn_check[btn].setChecked(True)
                break
        for i in [self.operation_zhengwen_radio, self.operation_dianfang_radio,
                  self.operation_genggai_radio, self.operation_hepiao_radio, self.operation_fenpaio_radio]:
            if i.text() == data_button[1]:
                i.setChecked(True)
                break
        for i in [self.operation_FOB_radio, self.operation_FIO_radio, self.operation_FILO_radio,
                  self.operation_FLT_radio]:
            if i.text() == data_button[2]:
                i.setChecked(True)
                break

        for i in range(0, len(self.list1)):

            if i in self.object_name_list:
                self.list1[i].disconnect()
                self.list1[i].setText(data[i])
                self.list1[i].installEventFilter(self)
                self.list1[i].textChanged.connect(lambda: self.change_listwidget(self.list1[i].objectName()))
                if isinstance(i, QLineEdit):
                    self.list1[i].editingFinished.connect(lambda: self.write_file(self.list1[i].objectName()))
            else:
                self.list1[i].setText(data[i])
        self.conn(self.kehumingcheng_lineedit)
        self.conn(self.kehucaozuo_lineedit)
        self.conn(self.baoguangongsi_lineedit)
        self.conn(self.dimiangongsi_lineedit)
        self.conn(self.bangzagongsi_lineedit)
        self.conn(self.chuandongmingcheng_lineedit)
        self.conn(self.huoming_lineedit)
        self.conn(self.jianshu_lineedit)
        self.conn(self.zhongliang_lineedit)
        self.conn(self.tiji_lineedit)
        self.conn(self.jifeidun_lineedit)
        self.conn(self.fahuoren_textedit)
        self.conn(self.shouhuoren_textedit)
        self.conn(self.tongzhiren_textedit)
        self.conn(self.maitou_textedit)
        self.conn(self.sheepname_lineedit)
        self.conn(self.chuandaigongsi_lineedit)
        self.conn(self.lianxiren_lineedit)
        self.conn(self.fahuoren_textedit)
        self.conn(self.tongzhiren_textedit)
        self.conn(self.huwuleixing_lineedit)

    def reset_lk(self):
        list = [self.bangzagongsi_lineedit, self.bangzha_teshuyaoqiu_lineedit, self.baoguangongsi_lineedit,
                self.chuandaigongsi_lineedit, self.chuandongmingcheng_lineedit, self.dimiangongsi_lineedit,
                self.huwuleixing_lineedit, self.kehucaozuo_lineedit, self.kehumingcheng_lineedit,
                self.lianxiren_lineedit, self.teshuyaoqiu_lineedit, self.IMOhao_lineedit, self.chuanci_lineedit,
                self.mudigangETA_lineedit, self.mudigang_lineedit,
                self.qiyungang_lineedit, self.qiyungangATB_lineedit, self.qiyungangATD_lineedit,
                self.qiyungangETA_lineedit, self.sheepname_lineedit, self.tidanhao_lineedit,
                self.fahuoren_textedit, self.huomiao_textedit, self.huoming_lineedit, self.jianshu_lineedit,
                self.jifeidun_lineedit,
                self.maitou_textedit, self.shouhuoren_textedit, self.tiji_lineedit, self.tongzhiren_textedit,
                self.zhongliang_lineedit,
                self.beizhu_lineedit]
        for i in list:
            if i in self.object_name_list:
                i.disconnect()
                i.setText("")
                i.installEventFilter(self)
                i.textChanged.connect(lambda: self.change_listwidget(i.objectName()))
                if isinstance(i, QLineEdit):
                    i.editingFinished.connect(lambda: self.write_file(i.objectName()))
            else:
                i.setText("")
                # 将btn默认不选中
        all_btn = [self.operation_fahuo_checkbox, self.operation_daohuo_checkbox,
                   self.operation_dizai_checkbox, self.checkBox_18, self.operation_jiancha_checkbox,
                   self.operation_fangxing_checkbox, self.operation_zhengwen_radio, self.operation_dianfang_radio,
                   self.operation_genggai_radio, self.operation_hepiao_radio, self.operation_fenpaio_radio,
                   self.operation_FOB_radio, self.operation_FIO_radio, self.operation_FILO_radio,
                   self.operation_FLT_radio
                   ]
        for i in all_btn:
            i.setCheckable(False)
            i.setCheckable(True)

    def SQL_lk(self):
        if not self.tidanhao_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>提单号为空</strong>")
            return None
        if not self.sheepname_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>船名为空</strong>")
            return None
        if self.sheepname_lineedit.text() == self.tidanhao_lineedit.text():
            QMessageBox.warning(self, "警告", "SQL 失败: <strong>船名不应该和提单号相同</strong>")
            return None
        data = []
        for i in self.list1:
            try:
                if isinstance(i, QLineEdit):
                    data.append(i.text())
                elif isinstance(i, QTextEdit):
                    data.append(i.toPlainText())
            except:
                data.append("")
        data1 = data[:17]
        data2 = data[17:]
        btn_check = [self.operation_fahuo_checkbox, self.operation_daohuo_checkbox,
                     self.operation_dizai_checkbox, self.checkBox_18, self.operation_jiancha_checkbox,
                     self.operation_fangxing_checkbox]
        check_btn_data = [""]
        radio_one_data = [""]
        radio_two_data = [""]
        for i in range(len(btn_check)):
            if btn_check[i].isChecked():
                check_btn_data.append(btn_check[i].text())
            else:
                if i != 0:
                    check_btn_data = [check_btn_data[-1]]
                    break
                else:
                    check_btn_data.append("")
        for i in [self.operation_zhengwen_radio, self.operation_dianfang_radio,
                  self.operation_genggai_radio, self.operation_hepiao_radio, self.operation_fenpaio_radio]:
            if i.isChecked():
                radio_one_data.append(i.text())
                break
        for i in [self.operation_FOB_radio, self.operation_FIO_radio, self.operation_FILO_radio,
                  self.operation_FLT_radio]:
            if i.isChecked():
                radio_two_data.append(i.text())
                break
        data = data1 + [check_btn_data[-1]] + [radio_one_data[-1]] + [radio_two_data[-1]] + data2
        data.append(self.user)
        Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).save_sanhuo_oper(data=data, tidanhao=self.text)

    def help_lk(self):
        self.help_content = '<p style="line-height: 25px; font-size: 16px" ><b>下货纸\费用单 :</b>下货纸导出文件位置是当前软件同级目录下的Excel_files,' \
                            '如果需要修改模板样式，请到template文件夹下修改'',且不要修改模板中多次出现的数据<br>''<p style="line-height: 25px; font-size: ' \
                            '16px" ><b>刷新: </b>点击刷新会重新从数据库提取数据中所用拥有的数据到当前界面<br>''<p style="line-height: 25px; font-size: ' \
                            '16px" ><b>SQL: </b>会将当前数据保存到数据库<br>''<p style="line-height: 25px; font-size: 16px" ><b>重置: ' \
                            '</b>点击重置会清空当前界面输入框所有内容<br>''<p style="line-height: 25px; font-size: 16px" ><b>其他: ' \
                            '</b>向上滑动滚轮界面将上移,向下滑动滚轮界面将下移<br> ' \
                            '<p style="line-height: 25px; font-size: 16px"; font-color:red ><b>注意 ' \
                            ':</b>由于软件需要,请操作完后点击退出保证软件正常运行'

        QMessageBox.information(self, '帮助', self.help_content)

    def conn(self, object_name):
        """

        :param object_name: lineedit对象
        :param listwidget: listwidget对象
        :return: none
        """
        self.object_name_list.append(object_name)
        object_name.installEventFilter(self)
        self.listwidget.hide()
        self.listwidget.clicked.connect(self.lineedit_set_text)

        object_name.textChanged.connect(lambda: self.change_listwidget(object_name.objectName()))
        if isinstance(object_name, QLineEdit):
            object_name.editingFinished.connect(lambda: self.write_file(object_name.objectName()))
        # 创建数据存储的文件
        self.listwidget_data = {}
        self.remember_data = {}
        row_data = []

        self.file = self.file = os.path.abspath(".") + "\\resource\\" + 'listwidget_data.xlsx'
        if os.path.exists(self.file):
            wb = openpyxl.load_workbook(filename=self.file)
            sheets = wb.sheetnames
            wb.close()
            if object_name.objectName() not in sheets:
                wb = openpyxl.load_workbook(filename=self.file)
                wb.create_sheet(object_name.objectName())
                wb.save(self.file)
                wb.close()
        else:
            wb = openpyxl.Workbook()
            wb.create_sheet(object_name.objectName())
            wb.save(self.file)
            wb.close()
        wb = openpyxl.load_workbook(filename=self.file, read_only=True)
        sheets = wb.sheetnames
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        row_data.append(cell.value)
            self.listwidget_data[sheet_name] = row_data
            self.remember_data[sheet_name] = []
            row_data = []
        self.current_listdata = []
        wb.close()

    def my_re(self, str):
        if isinstance(self.focuse_widget, QTextEdit):
            target = self.focuse_widget.toPlainText()
        else:
            target = self.focuse_widget.text()
        """
        :param traget: 目标字符串
        :param str: 查找的字符串
        :return: 前几位匹配的字符串
        """
        if re.match(target, str, re.I):
            return str

    #
    # 当lineedit控件输入框内容变化时触发，查找所有历史输入若存在首位相同显示listwidget控件
    # 并向中插入匹配的结果作为item.每次显示时清空上次items项
    def change_listwidget(self, objectname):
        objectname = self.focuse_widget.objectName()
        if self.listwidget_data.get(objectname):
            self.current_listdata = list(
                filter(lambda t: t is not None, list(map(self.my_re, self.listwidget_data.get(objectname)))))

            if self.current_listdata:
                self.listwidget.disconnect()
                self.listwidget.clear()
                self.listwidget.clicked.connect(self.lineedit_set_text)
                self.listwidget.addItems(sorted(self.current_listdata))
                self.listwidget.show()
        if isinstance(self.focuse_widget, QTextEdit):
            if self.focuse_widget.toPlainText() and not self.current_listdata:
                self.listwidget.hide()
        elif self.focuse_widget.text() and not self.current_listdata:

            self.listwidget.hide()

        if not self.listwidget.count():
            self.listwidget.hide()

    # 当点击listwidget的item时触发,将点击的item内容设置为当前lineedit文本
    def lineedit_set_text(self, index):
        row = index.row()
        self.focuse_widget.setText(self.current_listdata[row])

    # 编辑完成时触发，将lineedit文本内容与所有历史输入内容对比，不同则记住作为本次新增的输入内容
    def write_file(self, object):
        object = self.focuse_widget
        objectname = self.focuse_widget.objectName()
        if isinstance(object, QLineEdit):
            text = self.focuse_widget.text()
        else:
            text = object.toPlainText()
        try:
            if text in self.listwidget_data.get(objectname):
                pass
            else:
                if text:
                    self.remember_data[objectname].append(text)
                    self.listwidget_data[objectname].append(text)
        except:
            if text:
                self.remember_data[objectname].append(text)
                self.listwidget_data[objectname].append(text)

    # 当关闭程序时触发，将本次记住的新增的输入内容写入文件
    def closeEvent(self, event):
        showdata = ""
        count = 1  # 用来计数
        if "" in self.combobox_data:
            self.combobox_data = self.combobox_data[:-1]
        for i in self.combobox_data:
            try:
                if self.history_data.get(i):
                    count -= 1
                    pass
                else:
                    self.history_data[i] = ""
                    showdata += str(count) + ": " + i + "<br>"
            except:
                self.history_data[i] = ""
                showdata += str(count) + ": " + i + "<br>"
            count += 1
        if showdata:
            selectdata = [i[-3:] for i in showdata.split("<br>")[:-1]]
            selectdata.insert(0, {"tidanhao_num": self.tidanhao_lineedit.text()})
            res_data = Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).select_if_complate_data(selectdata, grade='2')
            if res_data != "all_exists":
                QMessageBox.about(self, '未完成操作', res_data if res_data else showdata)
        reply = QMessageBox.question(self, "提问", "确定退出吗?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.history_data["time"] = datetime.datetime.now().strftime("%Y-%m-%d-%H:%M.%S")
            self.remember_history.emit(self.user, 2, self.ship, self.text, self.tidanhao_lineedit.text(),
                                       self.combobox_data, self.history_data)
            # 将时间插入数据列表中
            Mysql(self.g_User, self.g_Pw, self.g_Port, self.g_Ip).save_sanhuo_time(self.text,
                                                                                   self.tidanhao_lineedit.text(),
                                                                                   self.combobox_data,
                                                                                   self.history_data)
            self.history_data = {}
            wb = openpyxl.load_workbook(self.file)
            sheets = wb.sheetnames
            for sheet_name in sheets:
                sheet = wb[sheet_name]
                if self.remember_data.get(sheet_name):
                    sheet.append(self.remember_data.get(sheet_name))
            wb.save(self.file)
            wb.close()

            event.accept()
        if reply == QMessageBox.No:
            event.ignore()

    #
    def eventFilter(self, a0: 'QObject', a1: 'QEvent'):
        if a1.type() == QEvent.FocusIn:

            self.focuse_widget = a0
            if isinstance(a0, QLineEdit):
                a0.setStyleSheet("border-bottom:none;")
            t = a0
            x = 0
            y = t.height()

            while True:
                x += t.x()
                y += t.y()
                t = t.parent()
                try:
                    check = t.parent().pos()
                except:
                    break
            self.listwidget.move(x, y)
            self.listwidget.disconnect()
            self.listwidget.clear()
            if self.listwidget_data.get(a0.objectName()):
                self.current_listdata = sorted(self.listwidget_data.get(a0.objectName()))
            else:
                self.current_listdata = []
                self.listwidget.hide()
            if self.current_listdata:
                self.listwidget.addItems(self.current_listdata)
            self.listwidget.pressed.connect(self.lineedit_set_text)
            if self.current_listdata:
                self.listwidget.show()
                self.listwidget.raise_()
            self.listwidget.setStyleSheet("QListWidget{"
                                          "border:1px solid rgba(30,30,30,0.3);"
                                          "background-color:rgb(230,230,230);"
                                          "}"
                                          "QListWidget::item"
                                          "{"
                                          "border:1px dashed gray;"
                                          "  }"

                                          "QScrollBar"
                                          "{"
                                          "width:px;"
                                          "background:rgb(0,0,0,0%);"
                                          "margin:0px,0px,0px,0px;"
                                          "padding-top:0px;"
                                          "padding-bottom:0px;"
                                          "}")
            self.listwidget.setFixedWidth(a0.width())
            self.listwidget.setFixedHeight(100)
        elif a1.type() == QEvent.FocusOut:
            #
            if isinstance(a0, QLineEdit):
                a0.setStyleSheet("border-bottom:1px solid rgb(0,0,0);")
            self.listwidget.move(2000, 2000)
            if isinstance(a0, QTextEdit):
                self.write_file(a0)
        return QWidget.eventFilter(self, a0, a1)

    def wheelEvent(self, event):
        if self.mapFromGlobal(self.cursor().pos()).x() in range(self.listwidget.pos().x(),
                                                                self.listwidget.pos().x() + self.listwidget.width()) and self.mapFromGlobal(
            self.cursor().pos()).y() in range(self.listwidget.pos().y(),
                                              self.listwidget.pos().y() + self.listwidget.height()) and self.listwidget.isVisible():
            pass
        elif QWheelEvent.angleDelta(event).y() > 0:
            self.move(self.pos().x(), self.pos().y() + 15)
        elif QWheelEvent.angleDelta(event).y() < 0:
            self.move(self.pos().x(), self.pos().y() - 15)

    def keyPressEvent(self, event):
        if event.key() == 16777220:
            try:
                if isinstance(self.focuse_widget, QLineEdit):
                    self.listwidget.hide()
            except:
                pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main = User_operation()
    main.show()
    app.exit(app.exec_())
